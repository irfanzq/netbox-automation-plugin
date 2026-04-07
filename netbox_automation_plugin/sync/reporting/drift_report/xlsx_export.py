"""Excel export for the drift audit (openpyxl).

Detail tables use the same ``HEADERS_*`` lists as ``format_html_proposed`` /
``drift_overrides_apply``. BMC sheets omit the ``MAAS NIC model`` column (NIC sheets still include
it). Rows are padded/truncated to header width; legacy BMC rows with an extra cell between
``MAAS BMC MAC`` and ``MAAS LLDP switch`` have that cell dropped so columns stay aligned.
"""

from io import BytesIO

from netbox_automation_plugin.sync.reporting.drift_report.constants import (
    _PHASE0_FIELD_OWNERSHIP_BULLETS,
    _PHASE0_FIELD_OWNERSHIP_LEAD,
    _PHASE0_FIELD_OWNERSHIP_TITLE,
)
from netbox_automation_plugin.sync.reporting.drift_report.fabric_alignment import (
    _alignment_review_rows,
)
from netbox_automation_plugin.sync.reporting.drift_report.metrics import (
    _count_hints,
    _phase0_category_counts,
    _run_metadata_rows,
    _severity_triage_rows,
)
from netbox_automation_plugin.sync.reporting.drift_report.placement import _drift_for_user_reports
from netbox_automation_plugin.sync.reporting.drift_report.drift_overrides_apply import (
    HEADERS_BMC_EXISTING,
    HEADERS_BMC_NEW_DEVICES,
    HEADERS_DETAIL_EXISTING_FIPS,
    HEADERS_DETAIL_EXISTING_PREFIXES,
    HEADERS_DETAIL_EXISTING_VMS,
    HEADERS_DETAIL_NEW_DEVICES,
    HEADERS_DETAIL_NEW_FIPS,
    HEADERS_DETAIL_NEW_IP_RANGES,
    HEADERS_DETAIL_NEW_NICS,
    HEADERS_DETAIL_NEW_PREFIXES,
    HEADERS_DETAIL_NEW_VMS,
    HEADERS_DETAIL_NIC_DRIFT,
    HEADERS_DETAIL_PROPOSED_MISSING_VLANS,
    HEADERS_PLACEMENT_ALIGNMENT,
    HEADERS_SERIAL_REVIEW,
    merge_drift_review_overrides,
    normalize_drift_review_overrides,
    _new_nic_row_is_os_authority,
    _update_nic_row_is_os_authority,
)
from netbox_automation_plugin.sync.reporting.drift_report.proposed_changes import (
    _proposed_changes_rows,
)


def _coerce_bmc_row_to_headers(headers: list[str], row: list | tuple) -> list:
    """If a row has one extra value between MAAS BMC MAC and MAAS LLDP switch, drop it (legacy)."""
    rl = list(row) if isinstance(row, (list, tuple)) else [row]
    n = len(headers)
    if len(rl) != n + 1:
        return rl
    try:
        i_mac = headers.index("MAAS BMC MAC")
        i_lldp = headers.index("MAAS LLDP switch")
    except ValueError:
        return rl
    if i_lldp != i_mac + 1 or len(rl) <= i_mac + 1:
        return rl
    del rl[i_mac + 1]
    return rl


def build_drift_report_xlsx(
    maas_data,
    netbox_data,
    openstack_data,
    drift,
    *,
    matched_rows=None,
    os_subnet_hints=None,
    os_subnet_gaps=None,
    os_ip_range_gaps=None,
    os_floating_gaps=None,
    netbox_prefix_count=0,
    netbox_inventory_counts=None,
    interface_audit=None,
    netbox_ifaces=None,
    drift_overrides=None,
):
    """
    Build an Excel (.xlsx) workbook from the same inputs as format_drift_report.
    Returns bytes suitable for HttpResponse(..., content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet").
    Google Sheets opens .xlsx files.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export. pip install openpyxl")

    orphaned_nb_count = len((drift or {}).get("in_netbox_not_maas") or [])
    drift = _drift_for_user_reports(drift)

    wb = Workbook()
    header_font = Font(bold=True)

    def _sheet(name, max_len=31):
        s = wb.create_sheet(title=name[:max_len])
        return s

    def _append_header(ws, row):
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(row) + 1):
            ws.cell(row=r, column=c).font = header_font
        return r

    # --- Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Drift audit summary"])
    ws_sum.cell(row=1, column=1).font = header_font
    ws_sum.append([])
    r_own = ws_sum.max_row + 1
    ws_sum.append([_PHASE0_FIELD_OWNERSHIP_TITLE, "", ""])
    ws_sum.cell(row=r_own, column=1).font = header_font
    ws_sum.append([_PHASE0_FIELD_OWNERSHIP_LEAD, "", ""])
    for b in _PHASE0_FIELD_OWNERSHIP_BULLETS:
        ws_sum.append([f"  • {b}", "", ""])
    ws_sum.append([])
    r_rm = ws_sum.max_row + 1
    ws_sum.append(["RUN METADATA", ""])
    ws_sum.cell(row=r_rm, column=1).font = header_font
    _append_header(ws_sum, ["Property", "Value"])
    for prop, val in _run_metadata_rows(maas_data, netbox_data, openstack_data):
        ws_sum.append([prop, val])
    ws_sum.append([])
    scope_meta = (drift or {}).get("scope_meta") or {}
    nb_devices_included = len(netbox_data.get("devices") or [])
    nb_devices_fetched = int(scope_meta.get("netbox_devices_before") or nb_devices_included)
    nb_sites_fetched = len(netbox_data.get("sites") or [])
    nb_sites_included = len({
        (d.get("site_slug") or "").strip()
        for d in (netbox_data.get("devices") or [])
        if (d.get("site_slug") or "").strip()
    })
    ws_sum.append(["MAAS", "OK" if not maas_data.get("error") else "Error", ""])
    ws_sum.append(["  Machines", str(len(maas_data.get("machines") or [])), ""])
    ws_sum.append(["NetBox", "OK" if not netbox_data.get("error") else "Error", ""])
    ws_sum.append(["  Devices (included / fetched)", f"{nb_devices_included} / {nb_devices_fetched}", ""])
    ws_sum.append(["  Sites (included / fetched)", f"{nb_sites_included} / {nb_sites_fetched}", ""])
    if netbox_prefix_count:
        ws_sum.append(["  IPAM Prefixes (included / fetched)", f"{netbox_prefix_count} / {netbox_prefix_count}", ""])
    nic = netbox_inventory_counts or {}
    if nic:
        ws_sum.append(["  Virtual machines (total)", str(nic.get("virtual_machines", 0)), ""])
        ws_sum.append(["  IP addresses (total)", str(nic.get("ip_addresses_total", 0)), ""])
        ws_sum.append(["  IP addresses VIP role (FIP-style)", str(nic.get("ip_addresses_vip_role", 0)), ""])
        ws_sum.append(
            [
                "  IP addresses with NAT inside set (outside/public side)",
                str(nic.get("ip_addresses_nat_outside", 0)),
                "",
            ]
        )
    ws_sum.append([])
    ws_sum.append([])
    pc = _phase0_category_counts(
        drift,
        matched_rows,
        interface_audit,
        os_subnet_gaps,
        os_floating_gaps or [],
    )
    serial_validation_needed = _count_hints(matched_rows, "NB serial empty")
    bmc_oob_mismatch = _count_hints(matched_rows, "BMC ")
    sub_txt = str(pc["sub_gaps"]) if pc["sub_gaps"] is not None else "N/A"
    if scope_meta:
        ws_sum.append([])
        ws_sum.append(["SCOPE", "", ""])
        _append_header(ws_sum, ["Check", "Value"])
        ws_sum.append(["Coverage status", str(scope_meta.get("coverage_status") or "PARTIAL")])
        ws_sum.append(["Selected sites", ", ".join(scope_meta.get("selected_sites") or []) or "(all)"])
        ws_sum.append(["Selected locations", ", ".join(scope_meta.get("selected_locations") or []) or "(all)"])
        ws_sum.append([
            "MAAS machines included / fetched",
            f"{scope_meta.get('maas_machines_after', 0)} / {scope_meta.get('maas_machines_before', 0)}",
        ])
        ws_sum.append([
            "NetBox devices included / fetched",
            f"{scope_meta.get('netbox_devices_after', 0)} / {scope_meta.get('netbox_devices_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack nets included / fetched",
            f"{scope_meta.get('openstack_networks_after', 0)} / {scope_meta.get('openstack_networks_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack subnets included / fetched",
            f"{scope_meta.get('openstack_subnets_after', 0)} / {scope_meta.get('openstack_subnets_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack FIPs included / fetched",
            f"{scope_meta.get('openstack_fips_after', 0)} / {scope_meta.get('openstack_fips_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack runtime NIC rows",
            str(len((openstack_data or {}).get("runtime_nics") or [])),
        ])
        ws_sum.append([
            "OpenStack runtime BMC rows",
            str(len((openstack_data or {}).get("runtime_bmc") or [])),
        ])
    ws_sum.append([])
    ws_sum.append(["DRIFT COUNTS", "", ""])
    _append_header(ws_sum, ["Category", "Count"])
    ws_sum.append(["In MAAS only (not in NetBox)", str(pc["maas_only"])])
    _outside_scope = (drift or {}).get("maas_in_netbox_outside_scope") or []
    _outside_n = len(_outside_scope)
    ws_sum.append(
        [
            "In MAAS scope but already in NetBox under another site/location (detail on Summary)",
            str(_outside_n),
        ]
    )
    ws_sum.append(
        [
            "Orphaned NetBox devices (not seen in MAAS this run; read-only here; tagging/cleanup deferred to a separate UI workflow because NetBox update sources include netbox-agent, scripts, and manual entries, not just MAAS)",
            str(orphaned_nb_count),
        ]
    )
    ws_sum.append(["Matched — review hints", str(pc["check_hosts"])])
    ws_sum.append(["NetBox serial missing", str(serial_validation_needed)])
    ws_sum.append(["NIC rows not OK", str(pc["iface_not_ok"])])
    ws_sum.append(["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])])
    ws_sum.append(["VLAN mismatch (runtime authority vs NetBox)", str(pc["vlan_drift_nic"])])
    ws_sum.append(["VLAN unverified from MAAS fallback", str(pc["vlan_unverified_nic"])])
    ws_sum.append(["OpenStack subnet → no Prefix", sub_txt])
    ws_sum.append(["OpenStack FIP → no IP record", str(pc["fip_gaps"])])
    ws_sum.append(["BMC vs NetBox OOB differs (OS/MAAS fallback)", str(bmc_oob_mismatch)])
    ws_sum.append([])
    ws_sum.append(["SEVERITY TRIAGE (why these matter)", "", ""])
    _append_header(ws_sum, ["Severity", "Category", "Count", "Why this matters"])
    for row in _severity_triage_rows(
        pc,
        serial_validation_needed=serial_validation_needed,
        bmc_oob_mismatch=bmc_oob_mismatch,
        netbox_outside_scope=_outside_n,
    ):
        ws_sum.append(row)
    ws_sum.append([])
    ws_sum.append(["RUN METRICS", "", ""])
    _append_header(ws_sum, ["Metric", "Value"])
    os_nics_included = len((openstack_data or {}).get("runtime_nics") or [])
    os_nics_fetched = int(scope_meta.get("openstack_runtime_nics_before") or os_nics_included)
    os_bmc_included = len((openstack_data or {}).get("runtime_bmc") or [])
    os_bmc_fetched = int(scope_meta.get("openstack_runtime_bmc_before") or os_bmc_included)
    ws_sum.append(["MAAS machines", str(len(maas_data.get("machines") or []))])
    ws_sum.append(["NetBox devices (included / fetched)", f"{nb_devices_included} / {nb_devices_fetched}"])
    ws_sum.append(["OpenStack runtime NIC rows (included / fetched)", f"{os_nics_included} / {os_nics_fetched}"])
    ws_sum.append(["OpenStack runtime BMC rows (included / fetched)", f"{os_bmc_included} / {os_bmc_fetched}"])
    ws_sum.append(["Hosts present in both MAAS and NetBox", str(drift.get("matched_count", 0))])
    ws_sum.append(["In MAAS only", str(pc["maas_only"])])
    ws_sum.append(["NetBox serial missing", str(serial_validation_needed)])
    ws_sum.append(["OpenStack subnet gaps", sub_txt])
    ws_sum.append(["OpenStack FIP gaps", str(pc["fip_gaps"])])
    ws_sum.append(["VLAN mismatch NICs (OS/MAAS authority)", str(pc["vlan_drift_nic"])])
    ws_sum.append(["VLAN unverified NICs (MAAS fallback)", str(pc["vlan_unverified_nic"])])
    ws_sum.append(["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])])
    ws_sum.append([])
    if _outside_scope:
        r_os = ws_sum.max_row + 1
        ws_sum.append(["Detail — MAAS in scope, NetBox outside selected site/location", ""])
        ws_sum.cell(row=r_os, column=1).font = header_font
        ws_sum.append([
            "These devices exist in NetBox; the scoped device list omitted them. Not new-device candidates.",
        ])
        _append_header(
            ws_sum,
            ["Hostname", "NetBox region", "NetBox site", "NetBox location", "Note"],
        )
        for row in _outside_scope:
            ws_sum.append(list(row))
        ws_sum.append([])
    align_rows_x = _alignment_review_rows(matched_rows)
    prop = _proposed_changes_rows(
        maas_data,
        netbox_data,
        drift,
        interface_audit,
        matched_rows,
        os_subnet_gaps or [],
        os_ip_range_gaps or [],
        os_floating_gaps or [],
        openstack_data=openstack_data,
        netbox_ifaces=netbox_ifaces,
        os_subnet_hints=os_subnet_hints or [],
    )
    norm = normalize_drift_review_overrides(drift_overrides)
    if norm:
        prop, align_rows_x = merge_drift_review_overrides(prop, align_rows_x, norm)
    if align_rows_x:
        r_al = ws_sum.max_row + 1
        ws_sum.append(["Detail — placement & lifecycle alignment", ""])
        ws_sum.cell(row=r_al, column=1).font = header_font
        _ph = list(HEADERS_PLACEMENT_ALIGNMENT)
        _append_header(ws_sum, _ph)
        _pn = len(_ph)
        for row in align_rows_x:
            rl = list(row) if isinstance(row, (list, tuple)) else [row]
            if len(rl) < _pn:
                rl = rl + [""] * (_pn - len(rl))
            elif len(rl) > _pn:
                rl = rl[:_pn]
            ws_sum.append(rl)
        ws_sum.append([])

    ws_sum.append([])
    ws_sum.append(["PROPOSED CHANGES (read-only)", "", ""])
    _append_header(ws_sum, ["Bucket", "Count"])
    total_props_x = (
        len(prop["add_devices"])
        + len(prop["add_prefixes"])
        + len(prop.get("update_prefixes", []))
        + len(prop.get("add_ip_ranges", []))
        + len(prop["add_fips"])
        + len(prop.get("update_fips", []))
        + len(prop.get("add_openstack_vms", []))
        + len(prop.get("update_openstack_vms", []))
        + len(prop["update_nic"])
        + len(prop["add_nb_interfaces"])
        + len(prop["add_mgmt_iface"])
        + len(prop.get("add_mgmt_iface_new_devices", []))
        + len(prop["review_serial"])
        + len(prop.get("add_proposed_missing_vlans", []))
    )
    ws_sum.append(["New devices", str(len(prop["add_devices"]))])
    ws_sum.append(["Review-only MAAS-only hosts", str(len(prop.get("add_devices_review_only", [])))])
    ws_sum.append(
        ["Proposed missing VLANs (IPAM)", str(len(prop.get("add_proposed_missing_vlans", [])))]
    )
    ws_sum.append(["New prefixes", str(len(prop["add_prefixes"]))])
    ws_sum.append(["Existing prefixes (drift)", str(len(prop.get("update_prefixes", [])))])
    ws_sum.append(["New IP ranges (allocation pools)", str(len(prop.get("add_ip_ranges", [])))])
    ws_sum.append(["New floating IPs", str(len(prop["add_fips"]))])
    ws_sum.append(["Existing floating IPs (NAT drift)", str(len(prop.get("update_fips", [])))])
    ws_sum.append(["New VMs (OpenStack)", str(len(prop.get("add_openstack_vms", [])))])
    ws_sum.append(["Existing VMs (drift)", str(len(prop.get("update_openstack_vms", [])))])
    ws_sum.append(["New NICs", str(len(prop["add_nb_interfaces"]))])
    ws_sum.append(["NIC drift", str(len(prop["update_nic"]))])
    ws_sum.append(["BMC / OOB", str(len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", [])))])
    ws_sum.append(["Serials (review)", str(len(prop["review_serial"]))])
    ws_sum.append(["Total", str(total_props_x)])

    # Matched-host drift worksheet intentionally suppressed to match on-screen report.

    # --- Proposed changes (full list) ---
    ws_prop = _sheet("Proposed changes")
    ws_prop.append(["Drift detail — read-only; OpenStack runtime is authoritative where present, MAAS is fallback."])
    ws_prop.cell(row=1, column=1).font = header_font
    ws_prop.append([])
    _append_header(ws_prop, ["Section", "Count"])
    nic_drift_os = [r for r in prop["update_nic"] if _update_nic_row_is_os_authority(r)]
    nic_drift_maas = [r for r in prop["update_nic"] if not _update_nic_row_is_os_authority(r)]
    ws_prop.append(["New devices (MAAS fallback)", len(prop["add_devices"])])
    ws_prop.append(["Review-only MAAS-only hosts", len(prop.get("add_devices_review_only", []))])
    ws_prop.append(
        ["Proposed missing VLANs (IPAM)", len(prop.get("add_proposed_missing_vlans", []))]
    )
    ws_prop.append(["New prefixes (OpenStack authority)", len(prop["add_prefixes"])])
    ws_prop.append(["Existing prefixes (drift)", len(prop.get("update_prefixes", []))])
    ws_prop.append(["New IP ranges (OpenStack authority)", len(prop.get("add_ip_ranges", []))])
    ws_prop.append(["New floating IPs (OpenStack authority)", len(prop["add_fips"])])
    ws_prop.append(["Existing floating IPs (NAT drift)", len(prop.get("update_fips", []))])
    ws_prop.append(["New VMs (OpenStack)", len(prop.get("add_openstack_vms", []))])
    ws_prop.append(["Existing VMs (drift)", len(prop.get("update_openstack_vms", []))])
    ws_prop.append(["New NICs", len(prop["add_nb_interfaces"])])
    ws_prop.append(["NIC drift (OS runtime authority)", len(nic_drift_os)])
    ws_prop.append(["NIC drift (MAAS fallback authority)", len(nic_drift_maas)])
    ws_prop.append(["BMC / OOB", len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", []))])
    ws_prop.append(["Serials (review)", len(prop["review_serial"])])

    def _append_block(title, headers, rows, *, note=None):
        ws_prop.append([])
        ws_prop.append([title])
        ws_prop.cell(row=ws_prop.max_row, column=1).font = header_font
        if note:
            ws_prop.append([note])
        hdr_list = list(headers)
        _append_header(ws_prop, hdr_list)
        ncols = len(hdr_list)
        bmc_sheet = "BMC / OOB" in title
        for row in rows:
            rl = list(row) if isinstance(row, (list, tuple)) else [row]
            if bmc_sheet:
                rl = _coerce_bmc_row_to_headers(hdr_list, rl)
            if len(rl) < ncols:
                rl = rl + [""] * (ncols - len(rl))
            elif len(rl) > ncols:
                rl = rl[:ncols]
            ws_prop.append(rl)

    _append_block(
        "A) New devices",
        list(HEADERS_DETAIL_NEW_DEVICES),
        prop["add_devices"],
    )
    _append_block(
        "A) MAAS-only hosts (manual review required)",
        list(HEADERS_DETAIL_NEW_DEVICES),
        prop.get("add_devices_review_only", []),
    )
    _append_block(
        "A) Proposed missing VLANs (IPAM)",
        list(HEADERS_DETAIL_PROPOSED_MISSING_VLANS),
        prop.get("add_proposed_missing_vlans", []),
    )
    _append_block(
        "A) New prefixes",
        list(HEADERS_DETAIL_NEW_PREFIXES),
        prop["add_prefixes"],
        note=(
            "NB proposed status: reserved when no Neutron ports were counted on that subnet in this scan; "
            "active when at least one port was seen (role certainty is only in Role reason). "
            "NB proposed VRF is inferred from OpenStack naming signals (network/project/region text)."
        ),
    )
    _append_block(
        "A) Existing prefixes (OpenStack drift)",
        list(HEADERS_DETAIL_EXISTING_PREFIXES),
        prop.get("update_prefixes", []),
    )
    _append_block(
        "A) New IP ranges (allocation pools)",
        list(HEADERS_DETAIL_NEW_IP_RANGES),
        prop.get("add_ip_ranges", []),
    )
    _append_block(
        "A) New floating IPs",
        list(HEADERS_DETAIL_NEW_FIPS),
        prop["add_fips"],
    )
    _append_block(
        "A) Existing floating IPs (NAT drift)",
        list(HEADERS_DETAIL_EXISTING_FIPS),
        prop.get("update_fips", []),
    )
    _append_block(
        "A) New VMs (OpenStack Nova)",
        list(HEADERS_DETAIL_NEW_VMS),
        prop.get("add_openstack_vms", []),
    )
    _append_block(
        "A) Existing VMs (OpenStack drift)",
        list(HEADERS_DETAIL_EXISTING_VMS),
        prop.get("update_openstack_vms", []),
    )
    new_nics_os = [r for r in prop["add_nb_interfaces"] if _new_nic_row_is_os_authority(r)]
    new_nics_maas = [r for r in prop["add_nb_interfaces"] if not _new_nic_row_is_os_authority(r)]
    _append_block(
        "B) New NICs (OS authority)",
        list(HEADERS_DETAIL_NEW_NICS),
        new_nics_os,
    )
    _append_block(
        "B) New NICs (MAAS authority)",
        list(HEADERS_DETAIL_NEW_NICS),
        new_nics_maas,
    )
    _append_block(
        "B) NIC drift (OS runtime authority)",
        list(HEADERS_DETAIL_NIC_DRIFT),
        nic_drift_os,
    )
    _append_block(
        "B) NIC drift (MAAS fallback authority)",
        list(HEADERS_DETAIL_NIC_DRIFT),
        nic_drift_maas,
    )
    _append_block(
        "B) BMC / OOB",
        list(HEADERS_BMC_EXISTING),
        prop["add_mgmt_iface"],
    )
    _append_block(
        "B) New-device BMC / OOB interfaces",
        list(HEADERS_BMC_NEW_DEVICES),
        prop.get("add_mgmt_iface_new_devices", []),
    )
    _append_block(
        "C) Serials",
        list(HEADERS_SERIAL_REVIEW),
        prop["review_serial"],
    )

    # Improve readability for long text cells in Excel/Sheets.
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is None:
                    continue
                # Keep existing horizontal alignment if present, but always wrap long text.
                cur = cell.alignment or Alignment()
                cell.alignment = Alignment(
                    horizontal=cur.horizontal,
                    vertical=cur.vertical or "top",
                    text_rotation=cur.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cur.shrink_to_fit,
                    indent=cur.indent,
                )
        # Give common text-heavy columns more space so wrapped content is easier to scan.
        ws.column_dimensions["A"].width = max(float(ws.column_dimensions["A"].width or 0), 42.0)
        if ws.max_column >= 2:
            ws.column_dimensions["B"].width = max(float(ws.column_dimensions["B"].width or 0), 26.0)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
