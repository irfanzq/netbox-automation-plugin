"""
Generate human-readable drift audit report from MAAS, NetBox, and OpenStack data.
Default UI output uses Bootstrap HTML tables; set use_html=False for ASCII (+---+) plain text.
XLSX export via build_drift_report_xlsx() for download (openpyxl); Google Sheets opens .xlsx.
NetBox-only device hostnames (in_netbox_not_maas) are cleared for HTML/XLSX and drift audit cache.

Copy in this module distinguishes **host data NICs** (Ethernet MAC/IP/VLAN) from **BMC / OOB**
(IPMI, iDRAC, Redfish — baseboard management controllers, not “another NIC”). NetBox models OOB
as **device OOB IP** plus an optional **OOB port** marked management-only in NetBox,
which documents the management attachment — not the same as in-band NICs.
"""

from collections import defaultdict
from io import BytesIO
import difflib
import html
import json
import re
import textwrap

_MAX_MAAS_MISSING_ROWS = 500
_MAX_OS_NETWORKS = 40
_MAX_OS_SUBNET_HINTS = 60
_MAX_COL = 10000
_MAX_MATCHED_COL = 18
_MAX_NOTES_COL = 42
# Notes column: full text for scrollable HTML report (avoid truncation)
_NOTES_COL_MAX_WIDTH = 8000
# Matched-hosts style tables: full cell text per column (no mid-cell truncation)
_DYNAMIC_COL_CAP = 10000
# ASCII tables: cap column width and wrap long cells so one outlier row does not pad every row.
_ASCII_COL_WRAP_DEFAULT = 96
_ASCII_NOTES_COL_WRAP = 200

# MAAS product → NetBox device type: best-score match (index + narrow + score).
_DT_MATCH_MIN_SCORE = 220.0
_DT_MATCH_TIE_EPSILON = 24.0
# Above this many types per vendor, pre-filter candidates before difflib scoring.
_DT_MATCH_NARROW_MIN = 20

_PHASE0_FIELD_OWNERSHIP_TITLE = "Field ownership — Phase 0 drift audit"
_PHASE0_FIELD_OWNERSHIP_LEAD = (
    "NetBox is the canonical inventory model in this workflow. "
    "MAAS and OpenStack supply runtime signals for comparison and gap detection; "
    "proposed actions are review-gated suggestions to align NetBox and do not run automatically."
)
_PHASE0_FIELD_OWNERSHIP_BULLETS = (
    "Discovery scope in Phase 0: detect new device candidates and drift; do not auto-apply changes.",
    "MAAS data is used for host/NIC matching and drift visibility (MAC, IP, VLAN observations).",
    "OOB/BMC values from MAAS power data inform NetBox management documentation alignment.",
    "OpenStack subnet and floating IP state is used to detect NetBox/IPAM gaps.",
    "Explicit non-goals for this phase: no blind hostname renames and no deletes from this screen.",
    "Conflict handling: operational evidence is compared to model intent; NetBox remains authoritative after approved updates.",
)


def _drift_for_user_reports(drift):
    """
    Shallow copy of drift for HTML/XLSX output and session cache.
    Clears in_netbox_not_maas (NetBox-only hostnames are not shown in user reports).
    """
    d = dict(drift or {})
    d["in_netbox_not_maas"] = []
    return d


def _dedupe_keep_order(items):
    seen = set()
    out = []
    for raw in (items or []):
        s = (str(raw or "")).strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _format_inventory_list(
    items,
    *,
    noisy_regex=None,
    noisy_label="auto-generated",
    sample=20,
    max_named_display=35,
):
    vals = _dedupe_keep_order(items)
    if not vals:
        return "(none)"
    if not noisy_regex:
        if len(vals) <= max_named_display:
            return ", ".join(vals)
        show = vals[:max_named_display]
        more = len(vals) - len(show)
        return f"{', '.join(show)}, ... +{more} more unique"
    pat = re.compile(noisy_regex)
    noisy = [v for v in vals if pat.match(v)]
    named = [v for v in vals if not pat.match(v)]
    parts = []
    if named:
        show_n = named[:max_named_display]
        more_n = len(named) - len(show_n)
        txt_n = ", ".join(show_n)
        if more_n > 0:
            txt_n = f"{txt_n}, ... +{more_n} more named"
        parts.append(f"named ({len(named)} unique): {txt_n}")
    if noisy:
        show = noisy[:sample]
        more = len(noisy) - len(show)
        txt = ", ".join(show)
        if more > 0:
            txt = f"{txt}, ... +{more} more"
        parts.append(f"{noisy_label} ({len(noisy)} unique): {txt}")
    return " | ".join(parts) if parts else "(none)"


def _maas_machine_by_hostname(maas_data):
    out = {}
    for m in maas_data.get("machines") or []:
        h = (m.get("hostname") or "").strip()
        if h:
            out[h] = m
    return out


def _nb_tokens(s: str) -> set[str]:
    return {t for t in re.split(r"[^a-z0-9]+", (s or "").lower()) if t}


def _fabric_matches_location_name(fabric_name: str, loc_name: str) -> bool:
    fab = (fabric_name or "").strip().lower()
    loc_l = (loc_name or "").strip().lower()
    if not fab or not loc_l:
        return False
    fab_compact = re.sub(r"[^a-z0-9]+", "", fab)
    loc_compact = re.sub(r"[^a-z0-9]+", "", loc_l)
    fab_tokens = _nb_tokens(fabric_name)
    loc_tokens = _nb_tokens(loc_name)
    if loc_compact and (loc_compact in fab_compact or fab_compact in loc_compact):
        return True
    return any(len(t) >= 4 and t in fab_tokens for t in loc_tokens)


def _site_meta_for_slug(netbox_data, site_slug: str) -> dict:
    slug = (site_slug or "").strip()
    if not slug:
        return {}
    for s in netbox_data.get("sites") or []:
        if (s.get("slug") or "").strip() == slug:
            return s
    return {}


def _site_slug_for_location_name(netbox_data, location_name: str) -> str:
    loc_l = (location_name or "").strip().lower()
    if not loc_l:
        return ""
    for loc in netbox_data.get("locations") or []:
        nm = (loc.get("name") or "").strip().lower()
        if nm and nm == loc_l:
            return (loc.get("site_slug") or "").strip()
    return ""


def _derive_site_slug_from_maas(machine: dict, fabric_map: dict, pool_map: dict) -> str:
    fab = (machine.get("fabric_name") or "").strip()
    pool = (machine.get("pool_name") or "").strip()
    if fab and fab != "-":
        for k, v in (fabric_map or {}).items():
            if str(k).strip().lower() == fab.lower():
                return str(v).strip()
    if pool and pool != "-":
        for k, v in (pool_map or {}).items():
            if str(k).strip().lower() == pool.lower():
                return str(v).strip()
    return ""


def _suggest_nb_location_for_fabric(
    fabric_name: str, site_slug: str, netbox_data: dict
) -> str:
    fab = (fabric_name or "").strip()
    if not fab or fab == "-":
        return ""
    site_slug = (site_slug or "").strip()
    best = ""
    for loc in netbox_data.get("locations") or []:
        ss = (loc.get("site_slug") or "").strip()
        if site_slug and ss and ss != site_slug:
            continue
        nm = (loc.get("name") or "").strip()
        if nm and _fabric_matches_location_name(fab, nm):
            return nm
    return best


def _netbox_placement_from_maas_machine(
    machine: dict,
    netbox_data: dict,
    fabric_map: dict,
    pool_map: dict,
) -> tuple[str, str, str]:
    """NB region, site display name, and location for a MAAS-only host (Detail — new devices logic)."""
    site_slug = _derive_site_slug_from_maas(machine, fabric_map, pool_map)
    site_meta = _site_meta_for_slug(netbox_data, site_slug) if site_slug else {}
    nb_site = (
        (site_meta.get("name") or site_slug or "—").strip() if site_slug else "—"
    )
    nb_region = (site_meta.get("region_name") or "—") if site_slug else "—"
    nb_loc = (
        _suggest_nb_location_for_fabric(
            str(machine.get("fabric_name") or ""), site_slug, netbox_data
        )
        or "—"
    )
    if (not site_slug) and nb_loc not in {"", "—"}:
        site_slug = _site_slug_for_location_name(netbox_data, nb_loc)
        if site_slug:
            site_meta = _site_meta_for_slug(netbox_data, site_slug)
            nb_site = (site_meta.get("name") or site_slug or "—").strip()
            nb_region = site_meta.get("region_name") or "—"
    return nb_region, nb_site, nb_loc


def _maas_vendor_product(machine: dict) -> tuple[str, str]:
    hi = machine.get("hardware_info")
    if isinstance(hi, str) and hi.strip().startswith("{"):
        try:
            hi = json.loads(hi)
        except Exception:
            hi = None
    if not isinstance(hi, dict):
        hi = {}
    vendor = (
        str(machine.get("hardware_vendor") or hi.get("system_vendor") or hi.get("vendor") or "")
    ).strip()
    product = (
        str(
            hi.get("system_product")
            or machine.get("hardware_product")
            or machine.get("product_name")
            or hi.get("product_name")
            or hi.get("mainboard_product")
            or ""
        )
    ).strip()
    if vendor.lower() in ("unknown", "none", "n/a", ""):
        vendor = ""
    if product.lower() in ("unknown", "none", "n/a", ""):
        product = ""
    return vendor, product


def _maas_product_sku_tokens(product: str) -> list[str]:
    """
    Pull hardware SKU-like tokens from a MAAS product string so we can match NetBox
    models that only store the short code (e.g. ``R660`` from ``PowerEdge R660``).
    """
    p = (product or "").strip()
    if not p:
        return []
    found: set[str] = set()
    # Dell PowerEdge rack / common patterns
    for m in re.finditer(r"\bR\d{3,4}[A-Z]{0,4}\b", p, re.I):
        found.add(m.group(0))
    for m in re.finditer(r"\bC\d{3,4}[A-Z]{0,4}\b", p, re.I):
        found.add(m.group(0))
    for m in re.finditer(r"\bDL\d{3,4}[A-Z]?\b", p, re.I):
        found.add(m.group(0))
    for m in re.finditer(r"\bXR\d{3,4}\w*\b", p, re.I):
        found.add(m.group(0))
    # Broader alnum product codes (e.g. other vendors)
    for m in re.finditer(r"\b[A-Z]{1,4}\d{3,5}[A-Z0-9\-]{0,6}\b", p, re.I):
        found.add(m.group(0))
    return sorted(found, key=lambda t: (-len(t), t.lower()))


def _device_type_contains_sku_token(dt: dict, token: str) -> bool:
    """True if token appears as its own token in model, slug, or display (not R660 inside R660xs)."""
    if not token:
        return False
    pat = re.compile(
        rf"(?<![A-Za-z0-9]){re.escape(token)}(?![A-Za-z0-9])",
        re.I,
    )
    for key in ("model", "slug", "display"):
        s = (dt.get(key) or "").strip()
        if s and pat.search(s):
            return True
    return False


def _fold_device_type_text(s: str) -> str:
    """Lowercase, alnum tokens, single spaces — for similarity and containment."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", (s or "").lower())).strip()


def _norm_vendor_dtype(v: str) -> str:
    """Normalize manufacturer string for bucketing (must match MAAS vendor normalization)."""
    s = re.sub(r"[^a-z0-9]+", " ", str(v or "").lower()).strip()
    toks = [t for t in s.split() if t]
    while toks and toks[-1] in {
        "inc",
        "incorporated",
        "corp",
        "corporation",
        "ltd",
        "llc",
        "co",
        "company",
    }:
        toks.pop()
    return " ".join(toks)


def _build_device_type_match_index(types_list: list) -> dict[str, list[dict]]:
    """
    One-time index: normalized manufacturer -> pre-folded NetBox device type rows.
    Speeds up per-host matching (no repeated string folding; smaller candidate sets).
    """
    buckets: dict[str, list[dict]] = defaultdict(list)
    for dt in types_list or []:
        if not isinstance(dt, dict):
            continue
        man_raw = (dt.get("manufacturer") or "").strip()
        man_norm = _norm_vendor_dtype(man_raw)
        model_raw = (dt.get("model") or "").strip()
        fold_m = _fold_device_type_text(model_raw)
        fold_d = _fold_device_type_text(dt.get("display") or "")
        fold_s = _fold_device_type_text((dt.get("slug") or "").replace("-", " "))
        fold_c = _fold_device_type_text(f"{man_raw} {model_raw}")
        tok_m = frozenset(fold_m.split()) if fold_m else frozenset()
        tok_d = frozenset(fold_d.split()) if fold_d else frozenset()
        tok_c = frozenset(fold_c.split()) if fold_c else frozenset()
        blob = f"{fold_m}{fold_s}{fold_d}".replace(" ", "").replace("-", "")
        buckets[man_norm].append(
            {
                "dt": dt,
                "model_lower": model_raw.lower(),
                "fold_m": fold_m,
                "fold_d": fold_d,
                "fold_s": fold_s,
                "fold_c": fold_c,
                "tok_m": tok_m,
                "tok_d": tok_d,
                "tok_c": tok_c,
                "blob": blob,
            }
        )
    return dict(buckets)


def _narrow_dtype_entries(
    entries: list[dict],
    prod_fold: str,
    prod_nums: list[str],
    sku_tokens: list[str],
) -> list[dict]:
    """Cheap filter before SequenceMatcher when a vendor has many device types."""
    if len(entries) <= _DT_MATCH_NARROW_MIN:
        return entries
    if not prod_fold:
        return entries
    prod_word_toks = frozenset(prod_fold.split())
    narrowed: list[dict] = []
    sku_slice = sku_tokens[:8]
    for e in entries:
        fm = e["fold_m"]
        if prod_fold == fm or prod_fold in fm or (fm and fm in prod_fold):
            narrowed.append(e)
            continue
        if prod_word_toks & e["tok_m"] or prod_word_toks & e["tok_d"] or prod_word_toks & e["tok_c"]:
            narrowed.append(e)
            continue
        if prod_nums and any(n in e["blob"] for n in prod_nums):
            narrowed.append(e)
            continue
        dt = e["dt"]
        for t in sku_slice:
            if _device_type_contains_sku_token(dt, t):
                narrowed.append(e)
                break
    return narrowed if narrowed else entries


def _seq_ratio_scaled(prod_fold: str, cand: str, scale: float) -> float:
    if len(cand) < 1:
        return 0.0
    return difflib.SequenceMatcher(None, prod_fold, cand).ratio() * scale


def _score_maas_product_vs_dtype_entry(
    prod_fold: str,
    product_n: str,
    sku_tokens: list[str],
    e: dict,
) -> float:
    """Same scoring ideas as before, using pre-folded index entry (faster per host)."""
    if not prod_fold:
        return 0.0
    model = e["fold_m"]
    disp = e["fold_d"]
    slug = e["fold_s"]
    combined = e["fold_c"]
    dt = e["dt"]

    parts: list[float] = []

    if model == prod_fold:
        parts.append(1000.0)
    if model:
        if prod_fold in model:
            parts.append(520.0 + min(len(prod_fold), 96) * 1.5)
        elif model in prod_fold and len(model) >= 3:
            parts.append(480.0 + min(len(model), 96) * 1.5)

    seen: set[str] = set()
    for cand in (model, disp, slug, combined):
        if not cand or cand in seen:
            continue
        seen.add(cand)
        parts.append(_seq_ratio_scaled(prod_fold, cand, 520.0))

    pt = set(prod_fold.split())
    seen_tok_cands: set[str] = set()
    for cand in (model, disp, combined):
        if not cand or cand in seen_tok_cands:
            continue
        seen_tok_cands.add(cand)
        ct = set(cand.split())
        if not pt or not ct:
            continue
        j = len(pt & ct) / len(pt | ct)
        parts.append(j * 420.0)

    nums = [n for n in re.findall(r"\d{3,5}", product_n) if len(n) >= 3]
    if nums:
        blob = e["blob"]
        matched = sum(1 for n in nums if n in blob)
        if matched:
            parts.append(95.0 * matched / len(nums))
        elif model:
            parts.append(-75.0)

    for tok in sku_tokens[:6]:
        if _device_type_contains_sku_token(dt, tok):
            parts.append(210.0)
            break

    return max(parts) if parts else 0.0


def _resolve_device_type_display(
    vendor_raw: str,
    product_n: str,
    dtype_index: dict[str, list[dict]],
) -> str:
    """
    Pick NetBox device type display for MAAS vendor/product using a pre-built index.
    Logic: exact model+vendor, else scored match among (narrowed) vendor bucket.
    """
    vendor_raw = (vendor_raw or "").strip()
    product_n = (product_n or "").strip()
    if not vendor_raw or not product_n:
        return "—"
    vendor_l = _norm_vendor_dtype(vendor_raw)
    entries = dtype_index.get(vendor_l) or []
    if not entries:
        return "—"

    exact_dt = [e["dt"] for e in entries if e["model_lower"] == product_n.lower()]
    if len(exact_dt) == 1:
        return (exact_dt[0].get("display") or "").strip() or "—"
    if len(exact_dt) > 1:
        return (exact_dt[0].get("display") or "").strip() + f" (ambiguous ×{len(exact_dt)})"

    prod_fold = _fold_device_type_text(product_n)
    prod_nums = [n for n in re.findall(r"\d{3,5}", product_n) if len(n) >= 3]
    sku_tokens = _maas_product_sku_tokens(product_n)
    candidates = _narrow_dtype_entries(entries, prod_fold, prod_nums, sku_tokens)

    scored = [
        (e["dt"], _score_maas_product_vs_dtype_entry(prod_fold, product_n, sku_tokens, e))
        for e in candidates
    ]
    scored.sort(
        key=lambda x: (
            -x[1],
            -len((x[0].get("model") or "")),
            (x[0].get("display") or "").lower(),
        )
    )
    best_dt, best_s = scored[0]
    if best_s < _DT_MATCH_MIN_SCORE:
        return "—"
    close = [
        dt
        for dt, s in scored
        if s >= best_s - _DT_MATCH_TIE_EPSILON and s >= _DT_MATCH_MIN_SCORE
    ]
    if len(close) > 1:
        return (close[0].get("display") or "").strip() + f" (ambiguous ×{len(close)})"
    return (best_dt.get("display") or "").strip() or "—"


def _is_storage_host_role(slug: str, name: str) -> bool:
    s = (slug or "").strip().lower()
    n = (name or "").strip().lower()
    if s == "storage-host" or n == "storage host":
        return True
    if "storage" in s and "host" in s and s != "storage":
        return True
    if "storage" in n and "host" in n and n.replace(" ", "") != "storage":
        return True
    return False


def _match_netbox_role_from_hostname(hostname: str, netbox_data: dict) -> str:
    hn = (hostname or "").strip().lower()
    if not hn:
        return "—"
    roles = netbox_data.get("device_roles") or []
    if not roles:
        return "—"
    tokens = {t for t in re.split(r"[^a-z0-9]+", hn) if t}

    def role_display(r: dict) -> str:
        name = (r.get("name") or "").strip()
        slug = (r.get("slug") or "").strip()
        return name or slug or "—"

    def pick_cpu_host() -> str:
        exact_name = [r for r in roles if (r.get("name") or "").strip().lower() == "cpu host"]
        if exact_name:
            return role_display(exact_name[0])
        exact_slug = [r for r in roles if (r.get("slug") or "").strip().lower() == "cpu-host"]
        if exact_slug:
            return role_display(exact_slug[0])
        for r in roles:
            s = (r.get("slug") or "").strip().lower()
            n = (r.get("name") or "").strip().lower()
            if "cpu" in s and "host" in s:
                return role_display(r)
            if "cpu" in n and "host" in n:
                return role_display(r)
        return "—"

    if "gpu" in tokens:
        for r in roles:
            s = (r.get("slug") or "").strip().lower()
            n = (r.get("name") or "").strip().lower()
            if "gpu" in s and "host" in s:
                return role_display(r)
            if "gpu" in n and "host" in n:
                return role_display(r)
        return "—"

    # Storage: token "stor" (org naming) OR token "weka" (explicit request), and hostname contains "se-s".
    if ("stor" in tokens or "weka" in tokens) and "se-s" in hn:
        candidates = []
        for r in roles:
            s = (r.get("slug") or "").strip().lower()
            n = (r.get("name") or "").strip().lower()
            if "weka" in tokens:
                if _is_storage_host_role(s, n):
                    candidates.append(r)
                    continue
            if _is_storage_host_role(s, n):
                continue
            if s == "storage" or n == "storage":
                candidates.append(r)
            elif "storage" in s and "host" not in s:
                candidates.append(r)
            elif "storage" in n and "host" not in n:
                candidates.append(r)
        if len(candidates) == 1:
            return role_display(candidates[0])
        if len(candidates) > 1:
            return role_display(candidates[0]) + f" (ambiguous ×{len(candidates)})"
        return "—"

    if "cpu" in tokens or "osctrl" in tokens:
        return pick_cpu_host()

    return "—"


def _new_device_fabric_display(maas_fabric_raw, nb_location_raw) -> str:
    """
    For new-device table, only show human fabric names that match NB location.
    Generic fabric-#### values are suppressed.
    """
    loc = (nb_location_raw or "").strip()
    if not loc or loc == "—":
        return "—"
    fabrics = _split_maas_fabrics(maas_fabric_raw)
    if not fabrics:
        return "—"
    non_generic = [f for f in fabrics if not _is_generic_maas_fabric_name(f)]
    if not non_generic:
        return "—"
    loc_l = loc.lower()
    for f in non_generic:
        fl = f.lower()
        if loc_l in fl or fl in loc_l:
            return f
        ftoks = {t for t in re.split(r"[^a-z0-9]+", fl) if t}
        ltoks = {t for t in re.split(r"[^a-z0-9]+", loc_l) if t}
        if any(len(t) >= 4 and t in ftoks for t in ltoks):
            return f
    return "—"


_MAAS_NEW_DEVICE_UNSAFE_STATUSES = {
    "FAILED_COMMISSIONING",
    "COMMISSIONING",
    "TESTING",
    "RESCUE_MODE",
    "EXITING_RESCUE_MODE",
    "RELEASING",
    "DEPLOYING",
    "FAILED",
    "BROKEN",
}


def _norm_maas_status(raw: str) -> str:
    return re.sub(r"[\s\-]+", "_", str(raw or "").strip()).upper()


def _has_usable_maas_fabric(machine: dict) -> bool:
    fab = str(machine.get("fabric_name") or "").strip().lower()
    return bool(fab and fab not in {"-", "unknown", "n/a", "none", "null"})


def _new_device_candidate_policy(
    machine: dict,
    nic_count: int,
    *,
    vendor: str = "",
    product: str = "",
) -> tuple[bool, str, int]:
    """
    Candidate policy for "A) Add to NetBox / Detail — new devices".

    Returns:
      (is_candidate, note, sort_rank)
    Lower sort_rank comes first.
    """
    st = _norm_maas_status(machine.get("status_name") or machine.get("status"))
    has_fabric = _has_usable_maas_fabric(machine)
    has_identity = bool((vendor or "").strip() and (product or "").strip())

    weak_flags = []
    if nic_count == 0:
        weak_flags.append("0 NICs")
    if not has_fabric:
        weak_flags.append("no MAAS fabric")
    if not has_identity:
        weak_flags.append("incomplete identity")
    weak_note = ", ".join(weak_flags)

    if st in _MAAS_NEW_DEVICE_UNSAFE_STATUSES:
        return False, f"MAAS status {st} is transient/unsafe for inventory create", 90
    if st == "DEFAULT" and weak_flags:
        return False, f"MAAS status DEFAULT with weak data ({weak_note})", 91
    if weak_flags:
        return False, f"Weak discovery data ({weak_note})", 92

    rank = {
        "DEPLOYED": 0,
        "ACTIVE": 1,
        "READY": 2,
        "ALLOCATED": 3,
    }.get(st, 10)
    return True, "Candidate", rank


def _normalize_ascii_cell(s):
    return str(s).replace("\n", " ").replace("\r", "") if s is not None else ""


def _cell(s, w, *, truncate=True):
    s = _normalize_ascii_cell(s)
    if truncate and len(s) > w:
        s = s[: max(1, w - 2)] + ".."
    return s.ljust(w)


def _wrap_cell_lines(text: str, width: int) -> list[str]:
    """Word-wrap cell text to width; no truncation (long tokens split)."""
    if width < 1:
        width = 1
    t = _normalize_ascii_cell(text)
    if not t:
        return [""]
    lines = textwrap.wrap(
        t,
        width=width,
        break_long_words=True,
        break_on_hyphens=False,
        replace_whitespace=False,
    )
    return lines if lines else [""]


def _ascii_table(
    headers,
    rows,
    indent="  ",
    *,
    max_col=None,
    notes_col_idx=None,
    dynamic_columns=True,
    wrap_max_width=_ASCII_COL_WRAP_DEFAULT,
):
    """
    dynamic_columns: width per column from content, capped by wrap_max_width, with word wrap
    (avoids one long cell forcing huge padding on every row). Full text is kept on extra lines.
    After each logical row (including wrapped multi-line rows), a horizontal rule is printed so
    continuation lines are not confused with the next record.

    wrap_max_width: None disables wrap cap (legacy: single-line rows, wide columns, may truncate
    via _cell if over cap). Default 96 balances readability in <pre> without horizontal sprawl.
    """
    if not headers:
        return []
    n = len(headers)

    if dynamic_columns and wrap_max_width is not None:
        widths = []
        for i in range(n):
            w = max(len(_normalize_ascii_cell(headers[i])), 6)
            for r in rows:
                if i < len(r):
                    w = max(w, len(_normalize_ascii_cell(r[i])))
            cap = wrap_max_width
            if notes_col_idx is not None and i == notes_col_idx:
                cap = min(max(cap, _ASCII_NOTES_COL_WRAP), _NOTES_COL_MAX_WIDTH)
            widths.append(min(w, cap))

        def emit_wrapped_row(cells: list) -> list[str]:
            padded = list(cells[:n]) + [""] * (n - min(len(cells), n))
            col_lines = [_wrap_cell_lines(padded[i], widths[i]) for i in range(n)]
            h = max(len(cl) for cl in col_lines)
            col_lines = [cl + [""] * (h - len(cl)) for cl in col_lines]
            lines_out = []
            for li in range(h):
                parts = [col_lines[i][li].ljust(widths[i]) for i in range(n)]
                lines_out.append(
                    indent + "|" + "|".join(" " + parts[i] + " " for i in range(n)) + "|"
                )
            return lines_out

        sep = indent + "+" + "+".join("-" * (w + 2) for w in widths) + "+"
        out = [sep]
        out.extend(emit_wrapped_row(headers))
        out.append(sep)
        for r in rows:
            out.extend(emit_wrapped_row(r))
            out.append(sep)
        return out

    widths = []
    for i in range(n):
        if dynamic_columns:
            w = max(len(str(headers[i])), 6)
            for r in rows:
                if i < len(r):
                    cell = _normalize_ascii_cell(r[i])
                    w = max(w, len(cell))
            if notes_col_idx is not None and i == notes_col_idx:
                widths.append(min(w, _NOTES_COL_MAX_WIDTH))
            else:
                widths.append(min(w, _DYNAMIC_COL_CAP))
            continue
        if notes_col_idx is not None and i == notes_col_idx:
            w = max(len(str(headers[i])), 8)
            for r in rows:
                if i < len(r):
                    cell = _normalize_ascii_cell(r[i])
                    w = max(w, len(cell))
            widths.append(min(w, _NOTES_COL_MAX_WIDTH))
            continue
        cap = max_col if max_col is not None else _MAX_COL
        w = min(max(len(headers[i]), 5), cap)
        for r in rows:
            if i < len(r):
                w = min(max(w, min(len(str(r[i])), cap + 10)), cap)
        widths.append(w)
    sep = indent + "+" + "+".join("-" * (w + 2) for w in widths) + "+"
    out = [sep]
    out.append(
        indent + "|" + "|".join(" " + _cell(headers[i], widths[i]) + " " for i in range(n)) + "|"
    )
    out.append(sep)
    for r in rows:
        padded = list(r[:n]) + [""] * (n - min(len(r), n))
        out.append(
            indent + "|" + "|".join(" " + _cell(padded[i], widths[i]) + " " for i in range(n)) + "|"
        )
    out.append(sep)
    return out


def _banner(title, char="=", width=72):
    line = char * width
    return [line, f"  {title}", line]


def _html_cell_content(s) -> str:
    raw = _normalize_ascii_cell(s)
    return html.escape(raw, quote=False).replace("\n", "<br />")


def _html_col_is_mac(header) -> bool:
    """True for column headers that represent a MAC address (not substrings like 'machines')."""
    return re.search(r"(?i)\bMAC\b", str(header or "")) is not None


def _html_col_is_risk(header) -> bool:
    return str(header or "").strip().lower() == "risk"


def _html_col_is_ip(header) -> bool:
    """
    True for column headers representing IP addresses/lists.
    Matches labels like 'IP', 'IPs', 'OOB IP', 'MAAS BMC IP', etc.
    """
    h = str(header or "")
    return re.search(r"(?i)\bIP(?:s)?\b", h) is not None


def _html_th_class(header) -> str:
    h = str(header or "")
    base = "small align-bottom text-nowrap"
    if _html_col_is_mac(h):
        return f"{base} text-nowrap font-monospace"
    if _html_col_is_ip(h):
        return f"{base} text-nowrap"
    if _html_col_is_risk(h):
        return f"{base} text-nowrap"
    return base


def _html_td_class(header, col_idx, notes_col_idx=None) -> str:
    h = str(header or "")
    parts = []
    if _html_col_is_mac(h):
        parts.extend(["align-top", "text-nowrap", "font-monospace"])
    elif _html_col_is_ip(h):
        parts.extend(["align-top", "text-nowrap"])
    elif _html_col_is_risk(h):
        parts.extend(["align-top", "text-nowrap"])
    else:
        parts.extend(["align-top", "text-nowrap"])
    if notes_col_idx is not None and col_idx == notes_col_idx:
        parts.append("text-muted")
    return " ".join(parts)


def _html_table(headers, rows, *, notes_col_idx=None):
    if not headers:
        return ""
    n = len(headers)
    hdr_strs = [str(h) for h in headers]
    ths = "".join(
        f'<th scope="col" class="{_html_th_class(h)}">{_html_cell_content(h)}</th>'
        for h in hdr_strs
    )
    body_parts = []
    for r in rows:
        padded = list(r[:n]) + [""] * (n - min(len(r), n))
        tds = []
        for i, cell in enumerate(padded):
            h = hdr_strs[i] if i < len(hdr_strs) else ""
            cls = _html_td_class(h, i, notes_col_idx)
            tds.append(f'<td class="{cls}">{_html_cell_content(cell)}</td>')
        body_parts.append("<tr>" + "".join(tds) + "</tr>")
    return (
        '<div class="table-responsive mb-3">'
        '<table class="table table-sm table-bordered table-striped align-middle mb-0">'
        f'<thead class="table-light"><tr>{ths}</tr></thead><tbody>'
        + "".join(body_parts)
        + "</tbody></table></div>"
    )


class _DriftReportEmitter:
    __slots__ = ("_html", "_parts")

    def __init__(self, *, use_html: bool = True):
        self._html = use_html
        self._parts: list[str] = []

    def banner(self, title: str, char: str = "=") -> None:
        if self._html:
            major = char == "="
            cls = (
                "drift-rpt-title border-bottom border-2 pb-2 mb-3 mt-3 fw-bold text-body"
                if major
                else "drift-rpt-subtitle border-bottom pb-2 mb-2 mt-3 fw-semibold text-body"
            )
            self._parts.append(f'<div class="{cls}">{html.escape(title)}</div>')
        else:
            self._parts.extend(_banner(title, char))

    def spacer(self) -> None:
        if self._html:
            self._parts.append('<div class="drift-rpt-spacer mb-2" aria-hidden="true"></div>')
        else:
            self._parts.append("")

    def subtitle(self, text: str) -> None:
        t = (text or "").strip()
        if self._html:
            self._parts.append(
                f'<div class="text-body-secondary fw-semibold mb-1">{html.escape(t)}</div>'
            )
        else:
            self._parts.append(f"  {t}")

    def paragraph(self, text: str) -> None:
        t = (text or "").strip()
        if not t:
            return
        if self._html:
            self._parts.append(f'<p class="small text-body-secondary mb-2">{html.escape(t)}</p>')
        else:
            self._parts.append(f"  {t}")

    def error(self, message: str) -> None:
        if self._html:
            self._parts.append(
                '<div class="alert alert-danger py-2 px-3 small mb-2" role="alert">'
                f"{html.escape(str(message))}</div>"
            )
        else:
            self._parts.append(f"    Error: {message}")

    def line_total(self, text: str) -> None:
        t = (text or "").strip()
        if self._html:
            self._parts.append(f'<p class="small fw-semibold mb-2">{html.escape(t)}</p>')
        else:
            self._parts.append(f"  {t}")

    def table(self, headers, rows, **kw) -> None:
        if self._html:
            self._parts.append(
                _html_table(headers, rows, notes_col_idx=kw.get("notes_col_idx"))
            )
        else:
            self._parts.extend(_ascii_table(headers, rows, **kw))

    def phase0_field_ownership(self) -> None:
        """Explain source-of-truth and why MAAS/OS columns feed proposed NetBox actions."""
        if self._html:
            lis = "".join(
                f"<li>{html.escape(b)}</li>" for b in _PHASE0_FIELD_OWNERSHIP_BULLETS
            )
            self._parts.append(
                '<div class="alert alert-info border py-2 px-3 mb-3 small drift-rpt-field-ownership" role="note">'
                f'<div class="fw-semibold mb-1">{html.escape(_PHASE0_FIELD_OWNERSHIP_TITLE)}</div>'
                f'<p class="small text-body-secondary mb-2">{html.escape(_PHASE0_FIELD_OWNERSHIP_LEAD)}</p>'
                f'<ul class="mb-0 ps-3">{lis}</ul></div>'
            )
        else:
            self._parts.extend(_banner(_PHASE0_FIELD_OWNERSHIP_TITLE, "-"))
            self._parts.append(f"  {_PHASE0_FIELD_OWNERSHIP_LEAD}")
            self._parts.append("")
            for b in _PHASE0_FIELD_OWNERSHIP_BULLETS:
                self._parts.append(f"  • {b}")
            self._parts.append("")

    def render(self) -> str:
        if self._html:
            return (
                '<div class="drift-report-html" style="font-size:0.8125rem">'
                + "\n".join(self._parts)
                + "</div>"
            )
        return "\n".join(self._parts)


def _phase0_category_counts(
    drift,
    matched_rows,
    interface_audit,
    os_subnet_gaps,
    os_floating_gaps,
):
    maas_only = len(drift.get("in_maas_not_netbox") or [])
    check_hosts = sum(1 for r in (matched_rows or []) if r.get("place_match") == "CHECK")
    iface_not_ok = 0
    maas_nic_missing_nb = 0
    vlan_drift_nic = 0
    vlan_unverified_nic = 0
    for b in (interface_audit or {}).get("hosts") or []:
        for row in b.get("rows") or []:
            st = row.get("status") or ""
            if st != "OK":
                iface_not_ok += 1
            if st == "NOT_IN_NETBOX":
                maas_nic_missing_nb += 1
            if st.startswith("VLAN_DRIFT"):
                vlan_drift_nic += 1
            notes = row.get("notes") or ""
            if "VLAN unverified:" in notes:
                vlan_unverified_nic += 1
    if os_subnet_gaps is None:
        sub_gaps = None
    else:
        sub_gaps = len(os_subnet_gaps or [])
    fip_gaps = len(os_floating_gaps or [])
    return {
        "maas_only": maas_only,
        "check_hosts": check_hosts,
        "iface_not_ok": iface_not_ok,
        "maas_nic_missing_nb": maas_nic_missing_nb,
        "sub_gaps": sub_gaps,
        "fip_gaps": fip_gaps,
        "vlan_drift_nic": vlan_drift_nic,
        "vlan_unverified_nic": vlan_unverified_nic,
    }


def _matched_hosts_with_drift(matched_rows):
    """Rows that have review hints (place_match CHECK) so we show drifting hosts only."""
    if not matched_rows:
        return []
    return [
        r for r in matched_rows
        if r.get("place_match") == "CHECK" or (r.get("hints") or [])
    ]


def _count_hints(matched_rows, needle: str) -> int:
    c = 0
    for r in (matched_rows or []):
        for h in (r.get("hints") or []):
            if needle in (h or ""):
                c += 1
                break
    return c


# Substrings of hints from sync/reconciliation/audit_detail.py (placement / lifecycle only).
_ALIGNMENT_HINT_SUBSTRINGS = (
    "MAAS fabric vs NB location",
    "NB location empty — MAAS has fabric",
    "MAAS deployed / NB staged",
)


def _hint_is_placement_alignment(h: str) -> bool:
    t = (h or "").strip()
    return any(marker in t for marker in _ALIGNMENT_HINT_SUBSTRINGS)


def _is_generic_maas_fabric_name(name: str) -> bool:
    return re.fullmatch(r"(?i)fabric-\d+", (name or "").strip()) is not None


def _split_maas_fabrics(raw) -> list[str]:
    # MAAS fabric values may be a single name or a delimited list.
    txt = str(raw or "").strip()
    if not txt or txt == "—":
        return []
    parts = [p.strip() for p in re.split(r"[;,]", txt) if p.strip()]
    out = []
    seen = set()
    for p in parts:
        k = p.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def _select_alignment_fabric(maas_fabric_raw, nb_location_raw) -> str:
    """
    Prefer location-related MAAS fabric names (e.g. spruce-staging) and ignore
    generic fabric-#### labels when meaningful names exist.
    """
    fabrics = _split_maas_fabrics(maas_fabric_raw)
    if not fabrics:
        return "—"
    non_generic = [f for f in fabrics if not _is_generic_maas_fabric_name(f)]
    candidates = non_generic or fabrics
    nb_loc = (nb_location_raw or "").strip().lower()
    if nb_loc:
        for f in candidates:
            fl = f.lower()
            if nb_loc in fl or fl in nb_loc:
                return f
            tokens = [t for t in re.split(r"[-_\s/]+", fl) if t]
            if nb_loc in tokens:
                return f
    return candidates[0]


def _alignment_review_rows(matched_rows):
    """
    Matched hosts with placement/lifecycle hints only (not serial, NIC, or BMC/OOB —
    those have dedicated report tables).
    """
    out = []
    for r in matched_rows or []:
        hints = r.get("hints") or []
        align = [h for h in hints if _hint_is_placement_alignment(h)]
        if not align:
            continue
        joined = _dedupe_note_parts("; ".join(align))
        out.append(
            [
                r.get("hostname") or "",
                _select_alignment_fabric(r.get("maas_fabric"), r.get("netbox_location")),
                r.get("netbox_site") or "—",
                r.get("netbox_location") or "—",
                r.get("maas_status") or "-",
                r.get("netbox_status") or "-",
                joined or "—",
            ]
        )
    return sorted(out, key=lambda row: (row[0] or "").lower())


def _truncate_run_meta(text: str, max_len: int = 240) -> str:
    s = (text or "").strip()
    if not s:
        return ""
    if len(s) <= max_len:
        return s
    return s[: max_len - 1].rstrip() + "…"


def _run_metadata_rows(maas_data, netbox_data, openstack_data):
    """
    Trust/context lines for reviewers: source reachability, match policy, scope, action mode.
    Returns rows for Property / Value tables.
    """
    maas_data = maas_data or {}
    netbox_data = netbox_data or {}

    def _maas_line():
        err = maas_data.get("error")
        if err:
            return f"Failed — {_truncate_run_meta(str(err))}"
        return "Reachable / success"

    def _netbox_line():
        err = netbox_data.get("error")
        if err:
            return f"Failed — {_truncate_run_meta(str(err))}"
        return "Reachable / success"

    def _openstack_line():
        if openstack_data is None:
            return "Skipped — OpenStack not fetched for this run"
        osd = openstack_data or {}
        if osd.get("openstack_cred_missing"):
            return "Not configured — OpenStack credentials missing"
        err = osd.get("error")
        if err:
            return f"Failed — {_truncate_run_meta(str(err))}"
        return "Reachable / success"

    return [
        ["MAAS source", _maas_line()],
        ["OpenStack source", _openstack_line()],
        ["NetBox source", _netbox_line()],
        ["Match logic", "Hostname + NIC MAC"],
        ["Scope", "Phase 0 audit only (discovery + drift + proposed actions)"],
        [
            "Action mode",
            "Read-only from this screen: no NetBox write. Branch apply/merge is handled in a separate NetBox branch workflow.",
        ],
    ]


def _severity_triage_rows(pc, *, serial_validation_needed: int, bmc_oob_mismatch: int):
    """
    Severity policy for Phase 0 review.
    Returns rows: [severity, category, count, why].
    """
    sub_count = pc["sub_gaps"] if pc["sub_gaps"] is not None else "N/A"
    return [
        [
            "High",
            "OpenStack FIP → no IP record",
            str(pc["fip_gaps"]),
            "Routable addresses may exist without NetBox/IPAM tracking.",
        ],
        [
            "High",
            "OpenStack subnet → no Prefix",
            str(sub_count),
            "Subnet usage exists but design intent is missing from IPAM.",
        ],
        [
            "High",
            "VLAN mismatch (MAAS vs NetBox)",
            str(pc["vlan_drift_nic"]),
            "Observed VLAN differs from modeled intent; can impact active services.",
        ],
        [
            "High",
            "BMC vs NetBox OOB differs",
            str(bmc_oob_mismatch),
            "Out-of-band access may target the wrong management endpoint.",
        ],
        [
            "Medium",
            "NIC rows not OK",
            str(pc["iface_not_ok"]),
            "Interface-level deltas need review (IP/MAC/VLAN alignment).",
        ],
        [
            "Medium",
            "MAAS NIC missing in NetBox",
            str(pc["maas_nic_missing_nb"]),
            "Operational interfaces are present but not yet modeled.",
        ],
        [
            "Low",
            "Matched — review hints",
            str(pc["check_hosts"]),
            "Review host alignment/lifecycle hints (fabric vs location, lifecycle state mismatches). NIC/OOB drift is covered in dedicated detail tables.",
        ],
        [
            "Low",
            "NetBox serial missing",
            str(serial_validation_needed),
            "Asset metadata quality gap; usually non-blocking for connectivity.",
        ],
        [
            "Info",
            "VLAN unverified from MAAS",
            str(pc["vlan_unverified_nic"]),
            "Observation is incomplete; confirm before applying intent changes.",
        ],
        [
            "Info",
            "In MAAS only (not in NetBox)",
            str(pc["maas_only"]),
            "Discovery candidate count for onboarding review.",
        ],
    ]


def _device_by_name(netbox_data):
    out = {}
    for d in (netbox_data or {}).get("devices") or []:
        n = (d.get("name") or "").strip()
        if n:
            out[n] = d
    return out


def _dedupe_note_parts(text: str) -> str:
    if not text:
        return ""
    parts = [p.strip() for p in str(text).replace("\n", " ").split(";")]
    seen = set()
    ordered = []
    for p in parts:
        if not p:
            continue
        k = p.lower()
        if k in seen:
            continue
        seen.add(k)
        ordered.append(p)
    return "; ".join(ordered)


def _ip_address_host(ip_str: str) -> str:
    """Host portion of an IP or CIDR string, lowercased."""
    if not ip_str or ip_str == "—":
        return ""
    return str(ip_str).split("/", 1)[0].strip().lower()


# NetBox **port names** operators use for BMC/OOB (heuristic coverage); *-nic here is a label, not “host NIC”.
_MGMT_INTERFACE_NAME_HINTS = frozenset({
    "ipmi",
    "idrac",
    "bmc",
    "ilo",
    "imm",
    "xcc",
    "drac",
    "oob",
    "mgmt",
    "mgnt",
    "ipmi-nic",
    "bmc-nic",
})


def _netbox_iface_name_suggests_oob(name: str) -> bool:
    """
    True if the NetBox interface *name* looks like an OOB / BMC port.

    Names are operator-defined in NetBox (manual, import, etc.) — MAAS does not supply
    interface labels. We match exact aliases (e.g. ``idrac``) or **substring** hits so
    values like ``e0 - idrac`` still count as OOB-style (exact-set-only would miss those
    and wrongly flag ``IP_OTHER_IFACE``).
    """
    n = (name or "").strip().lower()
    if not n:
        return False
    if n in _MGMT_INTERFACE_NAME_HINTS:
        return True
    return any(hint in n for hint in _MGMT_INTERFACE_NAME_HINTS)


def _suggested_netbox_mgmt_interface_name(
    power_type: str,
    hardware_vendor: str | None = None,
    hardware_product: str | None = None,
) -> str:
    """
    MAAS-only hint when **creating** a new OOB port — ``ipmi`` or ``idrac`` only.

    - ``power_type`` containing ``redfish`` or ``idrac`` → ``idrac``.
    - ``power_type`` containing ``ipmi``: if vendor **or** product text contains
      ``dell`` (case-insensitive), suggest ``idrac`` (Dell BMC is iDRAC even when
      MAAS uses the generic IPMI driver); otherwise ``ipmi``.
    - Any other power type → ``ipmi``.

    Prefer NetBox’s existing port name when the BMC IP is already on an interface
    (``_oob_port_hint_column``).
    """
    pl = (power_type or "").lower()
    if "redfish" in pl or "idrac" in pl:
        return "idrac"
    if "ipmi" in pl:
        combined = (
            f"{(hardware_vendor or '').strip()} {(hardware_product or '').strip()}"
        ).lower()
        if "dell" in combined:
            return "idrac"
        return "ipmi"
    return "ipmi"


def _oob_port_hint_column(cov: str, nb_ifn: str, maas_when_no_nb_port: str) -> str:
    """
    Value for the 'NB OOB port (hint)' column: prefer the NetBox port name when the BMC IP
    is already documented on a port (MGMT_IFACE / IP_OTHER_IFACE).
    """
    n = (nb_ifn or "").strip()
    if n and n != "—" and cov in ("MGMT_IFACE", "IP_OTHER_IFACE"):
        return n
    return maas_when_no_nb_port


def _nb_iface_carrying_ip(nb_ifaces: list, target_ip: str):
    """First NetBox interface dict whose ips[] contains target (host match)."""
    bh = _ip_address_host(target_ip)
    if not bh:
        return None
    for iface in nb_ifaces or []:
        for ip in iface.get("ips") or []:
            if _ip_address_host(ip) == bh:
                return iface
    return None


def _meaningful_maas_power_type(pt: str) -> bool:
    """True if MAAS reports a concrete power driver (not empty / manual / unknown)."""
    p = (pt or "").strip().lower()
    if not p or p in ("—", "-", "manual", "unknown"):
        return False
    return True


def _netbox_bmc_ip_coverage(nb_ifaces: list, bmc_ip: str):
    """
    How NetBox documents the MAAS BMC IP (on an OOB-dedicated port / Interface, or elsewhere).
    Returns (code, port_name_or_emdash, short_note).
    """
    bh = _ip_address_host(bmc_ip)
    if not bh:
        return "NO_BMC_MAAS", "—", ""
    mgmt_name = ""
    any_name = ""
    for iface in nb_ifaces or []:
        iname = (iface.get("name") or "").strip().lower()
        is_mgmt_named = _netbox_iface_name_suggests_oob(iname)
        is_mgmt_flag = bool(iface.get("mgmt_only"))
        for ip in iface.get("ips") or []:
            if _ip_address_host(ip) != bh:
                continue
            disp = (iface.get("name") or "?").strip()
            if is_mgmt_flag or is_mgmt_named:
                mgmt_name = disp
            if not any_name:
                any_name = disp
            break
    if mgmt_name:
        return "MGMT_IFACE", mgmt_name, "BMC IP on OOB-style NetBox port"
    if any_name:
        return "IP_OTHER_IFACE", any_name, "BMC IP present; port name not typical for OOB"
    return "NO_IFACE_IP", "—", "No NetBox port carries this BMC IP"


def _build_proposed_mgmt_interface_rows(
    matched_rows,
    maas_by_hostname: dict,
    netbox_ifaces,
):
    """
    Matched hosts: **BMC / OOB** from MAAS power (IPMI, iDRAC, Redfish — not host data NICs).

    Compares MAAS BMC IP to NetBox device OOB and to NetBox **OOB ports**.
    Rows when power_type is set but BMC IP is missing from the MAAS API, or when OOB/BMC is not OK.
    Aligned BMC/OOB (status OK) is omitted — drift report lists issues only.
    """
    from netbox_automation_plugin.sync.reconciliation.audit_detail import (
        _normalize_mac,
    )

    nb_if = netbox_ifaces if isinstance(netbox_ifaces, dict) else {}
    out = []
    for r in matched_rows or []:
        h = (r.get("hostname") or "").strip()
        if not h:
            continue
        m = maas_by_hostname.get(h) or {}
        bmc = (m.get("bmc_ip") or "").strip()
        pt = (m.get("power_type") or "").strip() or "—"
        maas_oob_new = _suggested_netbox_mgmt_interface_name(
            pt, m.get("hardware_vendor"), m.get("hardware_product")
        )
        nb_oob = (r.get("netbox_oob") or "").strip()
        maas_mac = (m.get("bmc_mac") or "").strip()
        maas_vlan = (m.get("bmc_vlan") or "").strip()
        nb_list = nb_if.get(h) or []

        if not bmc:
            if not _meaningful_maas_power_type(pt):
                continue
            cov = "NO_BMC_IP_MAAS"
            nb_ifn = "—"
            nb_mgmt_mac = "—"
            status = "NO_BMC_IP"
            action = (
                "MAAS power_type is set but no BMC IP in machine API — grant admin API key, "
                "use op=power_parameters, or configure power_address in MAAS; then re-run audit."
            )
            if maas_mac or maas_vlan:
                action += f" MAAS hints: MAC={maas_mac or '—'} VLAN={maas_vlan or '—'}."
            risk = "High"
            out.append([
                h,
                "—",
                pt,
                maas_mac or "—",
                maas_oob_new,
                nb_oob or "—",
                cov,
                nb_ifn,
                nb_mgmt_mac,
                status,
                action,
                risk,
            ])
            continue

        cov, nb_ifn, cov_note = _netbox_bmc_ip_coverage(nb_list, bmc)
        oob_port_hint = _oob_port_hint_column(cov, nb_ifn, maas_oob_new)
        oob_match = bool(nb_oob) and _ip_address_host(nb_oob) == _ip_address_host(bmc)
        nb_detail = _nb_iface_carrying_ip(nb_list, bmc)
        nb_mgmt_mac = (nb_detail.get("mac") or "—") if nb_detail else "—"
        nb_mgmt_vid = (
            str(nb_detail.get("untagged_vlan_vid") or "—") if nb_detail else "—"
        )

        if oob_match and cov == "MGMT_IFACE":
            status = "OK"
            action = (
                "Device OOB + OOB port align with MAAS BMC; keep NetBox port name "
                f"'{nb_ifn}' (source of truth)"
            )
            risk = "None"
        elif oob_match and cov == "NO_IFACE_IP":
            status = "ADD_MGMT_IFACE"
            action = (
                f"Add NetBox OOB port '{maas_oob_new}' (management-only), assign BMC IP "
                f"{bmc}/<prefix> (OOB, not a host NIC)"
            )
            risk = "Medium"
        elif oob_match and cov == "IP_OTHER_IFACE":
            status = "REVIEW"
            action = (
                f"BMC IP on NetBox port '{nb_ifn}'; mark as OOB/management-only if needed — "
                "do not rename to match MAAS power_type; NetBox name is source of truth"
            )
            risk = "Low"
        elif not oob_match and cov == "MGMT_IFACE":
            status = "SET_OOB"
            action = (
                f"Set device OOB IP to {bmc} (MAAS BMC); BMC IP already on port '{nb_ifn}' "
                "(keep NetBox port name)"
            )
            risk = "Low"
        elif cov == "NO_IFACE_IP":
            status = "ADD_OOB_AND_MGMT"
            action = (
                f"Set device OOB IP to {bmc}; add OOB port '{maas_oob_new}' "
                "(management-only) with BMC IP"
            )
            risk = "Medium"
        else:
            status = "REVIEW"
            action = cov_note or "Align device OOB / OOB port / MAAS BMC"
            risk = "Medium"

        status_before_mac = status
        if maas_mac and nb_mgmt_mac and nb_mgmt_mac != "—":
            mm = _normalize_mac(maas_mac)
            nm = _normalize_mac(nb_mgmt_mac)
            if mm and nm and mm != nm:
                if status == "OK":
                    status = "REVIEW"
                    risk = "Medium"
                # Do not keep “everything aligns” copy when MACs disagree.
                if status_before_mac == "OK":
                    action = (
                        f"BMC MAC mismatch: MAAS {maas_mac} vs NetBox port {nb_mgmt_mac} "
                        f"on '{nb_ifn}'. OOB IP matches MAAS; confirm which MAC is correct."
                    )
                else:
                    action += (
                        f" MAC mismatch: MAAS BMC MAC {maas_mac} vs NetBox port {nb_mgmt_mac}."
                    )
        if maas_vlan and nb_mgmt_vid not in ("", "—", "None"):
            if maas_vlan.strip() != str(nb_mgmt_vid).strip():
                action += (
                    f" VLAN hint: MAAS {maas_vlan} vs NetBox untagged {nb_mgmt_vid} on BMC port."
                )
                if status == "OK":
                    status = "REVIEW"
                    risk = "Low"

        if str(status).strip().upper() == "OK":
            continue
        out.append([
            h,
            bmc,
            pt,
            maas_mac or "—",
            oob_port_hint,
            nb_oob or "—",
            cov,
            nb_ifn,
            nb_mgmt_mac,
            status,
            action,
            risk,
        ])
    return sorted(out, key=lambda x: (x[0] or "").lower())


def _build_add_nb_interface_rows(interface_audit):
    """
    MAAS NICs with a MAC that do not match any NetBox port on the device.
    Preview: proposed new NetBox ports (+ VLAN, IPs from MAAS).
    """
    out = []
    for b in (interface_audit or {}).get("hosts") or []:
        hn = (b.get("hostname") or "").strip()
        if not hn:
            continue
        nb_site = (b.get("nb_site") or "—").strip()
        nb_loc = (b.get("nb_location") or "—").strip()
        for row in b.get("rows") or []:
            if (row.get("status") or "") != "NOT_IN_NETBOX":
                continue
            maas_if = (row.get("maas_if") or "").strip() or "—"
            maas_fab = str(row.get("maas_fabric") or "—")
            mac = (row.get("maas_mac") or "").strip()
            ips = (row.get("maas_ips") or "—").strip()
            vlan = str(row.get("maas_vlan") or "—")
            props = (
                f"MAC {mac}; untagged VLAN {vlan} (from MAAS); "
                f"IPs: {ips}"
            )
            out.append([
                hn,
                nb_site,
                nb_loc,
                maas_if,
                maas_fab,
                mac,
                ips,
                vlan,
                (
                    maas_if
                    if maas_if != "—"
                    else (f"maas-nic-{mac.replace(':', '')[-6:]}" if mac else "maas-nic")
                ),
                props,
                "Medium",
            ])
    return sorted(out, key=lambda x: (x[0] or "").lower())


def _friendly_note(raw: str) -> str:
    note = _dedupe_note_parts(raw or "")
    parts = [p.strip() for p in note.split(";") if p.strip()]
    low = note.lower()
    # Keep IP-alignment rows focused on IP only; VLAN/MAC details belong to other sections.
    for p in parts:
        pl = p.lower()
        if "ip on maas not on nb iface:" in pl:
            return p
    for p in parts:
        pl = p.lower()
        if "ip" in pl and ("missing" in pl or "not on" in pl or "gap" in pl):
            return p
    # Drop MAC-only details from IP table and keep user-facing wording clear.
    non_mac = [
        p for p in parts
        if ("mac" not in p.lower() and "interface name" not in p.lower())
    ]
    for p in non_mac:
        pl = p.lower()
        if "ip" in pl:
            return p
    if "mac" in low:
        return "IP mismatch detected on this interface; review interface IP assignment."
    if "ip gap" in low:
        return "IP found in MAAS but missing on matching NetBox interface."
    if "vlan drift" in low:
        return "VLAN mismatch between MAAS and NetBox interface."
    if "vlan unverified" in low:
        return "MAAS did not return VLAN ID; verify VLAN manually."
    if "mac mismatch" in low:
        return "Matching interface name found, but MAC differs."
    return note or "Review interface data."


def _drift_table_status_is_ok_only(display_status: str) -> bool:
    """
    True when the drift Status column is only OK (after comma-split).
    e.g. "OK" or "OK, OK" → skip; "MISSING_NB_VLAN, OK" → False (still drifting).
    """
    parts = [p.strip().upper() for p in (display_status or "").split(",") if p.strip()]
    if not parts:
        return False
    return all(p == "OK" for p in parts)


def _proposed_changes_rows(
    maas_data,
    netbox_data,
    drift,
    interface_audit,
    matched_rows,
    os_subnet_gaps,
    os_floating_gaps,
    netbox_ifaces=None,
):
    """Build user-friendly proposed change buckets (preview only)."""
    try:
        from netbox_automation_plugin.sync.config import get_sync_config

        _sync = get_sync_config()
        _fabric_site_map = _sync.get("site_mapping_fabric") or {}
        _pool_site_map = _sync.get("site_mapping_pool") or {}
    except Exception:
        _fabric_site_map, _pool_site_map = {}, {}

    def _primary_mac_from_maas(ifaces):
        rows = [r for r in (ifaces or []) if str(r.get("mac") or "").strip()]
        if not rows:
            return "—"
        with_ip = [r for r in rows if r.get("ips")]
        pick = with_ip[0] if with_ip else rows[0]
        return str(pick.get("mac") or "—")

    def _new_device_nic_rows():
        out = []
        for h in sorted(drift.get("in_maas_not_netbox") or []):
            m = by_h.get(h, {})
            _, nb_site, nb_loc = _netbox_placement_from_maas_machine(
                m, netbox_data, _fabric_site_map, _pool_site_map
            )
            for r in (m.get("interfaces") or []):
                mac = str(r.get("mac") or "").strip().lower()
                if not mac:
                    continue
                maas_if = str(r.get("name") or "").strip() or "—"
                maas_fab = str(r.get("iface_fabric") or m.get("fabric_name") or "—")
                ips = ", ".join(r.get("ips") or []) or "—"
                vlan = str(r.get("vlan_vid") or "—")
                suggested_name = (
                    maas_if
                    if maas_if != "—"
                    else f"maas-nic-{mac.replace(':', '')[-6:]}"
                )
                props = f"MAC {mac}; untagged VLAN {vlan} (from MAAS); IPs: {ips}"
                out.append(
                    [
                        h,
                        nb_site,
                        nb_loc,
                        maas_if,
                        maas_fab,
                        mac,
                        ips,
                        vlan,
                        suggested_name,
                        props,
                        "Medium",
                    ]
                )
        return sorted(out, key=lambda x: (x[0] or "").lower())

    def _new_device_bmc_rows():
        out = []
        for h in sorted(drift.get("in_maas_not_netbox") or []):
            m = by_h.get(h, {})
            bmc_ip = str(m.get("bmc_ip") or "").strip()
            power_type = str(m.get("power_type") or "").strip()
            if not bmc_ip and not power_type:
                continue
            mgmt = _suggested_netbox_mgmt_interface_name(
                m.get("power_type"),
                m.get("hardware_vendor"),
                m.get("hardware_product"),
            )
            action = (
                f"Create OOB interface '{mgmt}' (management-only)"
                + (f"; set OOB/BMC IP {bmc_ip}" if bmc_ip else "")
            )
            out.append(
                [
                    h,
                    bmc_ip or "—",
                    power_type or "—",
                    str(m.get("bmc_mac") or "—"),
                    mgmt,
                    bmc_ip or "—",
                    action,
                    "Medium",
                ]
            )
        return sorted(out, key=lambda x: (x[0] or "").lower())

    by_h = _maas_machine_by_hostname(maas_data)
    _dtype_index = _build_device_type_match_index(netbox_data.get("device_types") or [])
    add_mgmt_iface = _build_proposed_mgmt_interface_rows(matched_rows, by_h, netbox_ifaces)
    add_mgmt_iface_new_devices = _new_device_bmc_rows()
    add_nb_interfaces = _build_add_nb_interface_rows(interface_audit) + _new_device_nic_rows()
    add_nb_interfaces = sorted(add_nb_interfaces, key=lambda x: (x[0] or "").lower())

    add_devices = []
    add_devices_review_only = []
    for h in sorted(drift.get("in_maas_not_netbox") or []):
        m = by_h.get(h, {})
        ifaces = m.get("interfaces") or []
        nic_count = sum(1 for r in ifaces if str(r.get("mac") or "").strip())
        primary_mac = _primary_mac_from_maas(ifaces)
        bmc_ip = str(m.get("bmc_ip") or "—")
        power_type = str(m.get("power_type") or "—")
        bmc_present = "Yes" if (bmc_ip not in {"", "—"} or power_type not in {"", "—"}) else "No"
        nb_region, nb_site, nb_loc = _netbox_placement_from_maas_machine(
            m, netbox_data, _fabric_site_map, _pool_site_map
        )
        mvendor, mproduct = _maas_vendor_product(m)
        nb_dtype = _resolve_device_type_display(mvendor, mproduct, _dtype_index)
        nb_role = _match_netbox_role_from_hostname(h, netbox_data)
        maas_fabric_disp = _new_device_fabric_display(str(m.get("fabric_name", "-")), nb_loc)
        is_candidate, note, status_rank = _new_device_candidate_policy(
            m, nic_count, vendor=mvendor, product=mproduct
        )
        row = [
            h,
            nb_region,
            nb_site,
            nb_loc,
            nb_dtype,
            nb_role,
            maas_fabric_disp,
            str(m.get("status_name", "-")),
            str(m.get("serial") or "—"),
            power_type,
            bmc_present,
            str(nic_count),
            primary_mac,
            ("maas-discovered" if is_candidate else "review-only"),
            (
                "Create device + ports"
                if is_candidate
                else f"Review only — not a safe NetBox add candidate ({note})"
            ),
        ]
        if is_candidate:
            add_devices.append((status_rank, h.lower(), row))
        else:
            add_devices_review_only.append((status_rank, h.lower(), row))

    add_devices = [r for _, _, r in sorted(add_devices, key=lambda x: (x[0], x[1]))]
    add_devices_review_only = [
        r for _, _, r in sorted(add_devices_review_only, key=lambda x: (x[0], x[1]))
    ]

    add_prefixes = []
    for g in (os_subnet_gaps or []):
        add_prefixes.append([
            g.get("cidr", ""),
            g.get("network_name", "-"),
            g.get("network_id", ""),
            "-",
            "Create Prefix",
        ])

    add_fips = []
    for g in (os_floating_gaps or []):
        add_fips.append([
            g.get("floating_ip", ""),
            g.get("fixed_ip_address", "-"),
            g.get("project_name") or g.get("project_id") or "-",
            "-",
            "Create IPAddress",
        ])

    update_nic = []
    for b in (interface_audit or {}).get("hosts") or []:
        hn = b.get("hostname", "")
        for row in b.get("rows") or []:
            st = str(row.get("status") or "").strip()
            notes = row.get("notes") or ""
            maas_vlan = str(row.get("maas_vlan") or "—")
            nb_vlan = str(row.get("nb_vlan") or "—")

            # NOT_IN_NETBOX: dedicated "create interface" table (add_nb_interfaces), not NIC drift.
            if st == "NOT_IN_NETBOX":
                continue

            # Drift-only: aligned interfaces are not listed (case-insensitive OK).
            if st.upper() == "OK":
                continue

            statuses = []
            reasons = []
            actions = []
            risk = "Medium"

            if "VLAN_DRIFT" in st:
                if nb_vlan in {"", "—", "None", "none"}:
                    statuses.append("MISSING_NB_VLAN")
                    reasons.append("NetBox VLAN missing; MAAS VLAN present")
                    actions.append("Set NetBox untagged VLAN from MAAS VLAN")
                else:
                    statuses.append("VLAN_MISMATCH")
                    reasons.append("NetBox VLAN differs from MAAS VLAN")
                    actions.append("Change NetBox untagged VLAN to match MAAS VLAN")
                risk = "High"

            if "IP_GAP" in st:
                statuses.append("MISSING_NB_IP")
                reasons.append(_friendly_note(notes))
                actions.append("Add missing IP on NetBox port")

            note_l = notes.lower()
            if ("netbox mac empty" in note_l) or ("mac mismatch" in note_l):
                if "mac mismatch" in note_l:
                    statuses.append("MAC_MISMATCH")
                else:
                    statuses.append("MISSING_NB_MAC")
                reasons.append("NetBox MAC missing or mismatched")
                actions.append("Set NetBox port MAC from MAAS for reliable matching")

            if not statuses:
                statuses.append(st)
                reasons.append(_dedupe_note_parts(notes) or "Port review needed")
                actions.append("Review port alignment manually")

            status_cell = ", ".join(dict.fromkeys(statuses))
            if _drift_table_status_is_ok_only(status_cell):
                continue

            update_nic.append([
                hn,
                row.get("maas_if") or "",
                str(row.get("maas_fabric") or "—"),
                row.get("maas_mac") or "",
                row.get("maas_ips") or "",
                row.get("nb_if") or "—",
                row.get("nb_mac") or "—",
                row.get("nb_ips") or "—",
                maas_vlan,
                nb_vlan,
                status_cell,
                "; ".join(dict.fromkeys([r for r in reasons if r])),
                "; ".join(dict.fromkeys([a for a in actions if a])),
                risk,
            ])

    review_serial = []
    for r in (matched_rows or []):
        if any("NB serial empty" in (h or "") for h in (r.get("hints") or [])):
            review_serial.append([
                r.get("hostname", ""),
                str(r.get("maas_serial", "")),
                str(r.get("netbox_serial", "")),
                "Manual validation",
                "High",
            ])

    return {
        "add_devices": add_devices,
        "add_devices_review_only": add_devices_review_only,
        "add_prefixes": add_prefixes,
        "add_fips": add_fips,
        "update_nic": update_nic,
        "add_nb_interfaces": add_nb_interfaces,
        "add_mgmt_iface": add_mgmt_iface,
        "add_mgmt_iface_new_devices": add_mgmt_iface_new_devices,
        "review_serial": review_serial,
    }


def format_drift_report(
    maas_data,
    netbox_data,
    openstack_data,
    drift,
    *,
    matched_rows=None,
    os_subnet_hints=None,
    os_subnet_gaps=None,
    os_floating_gaps=None,
    netbox_prefix_count=0,
    interface_audit=None,
    netbox_ifaces=None,
    use_html=True,
):
    """
    Return {"drift": str, "reference": str, "drift_markup": "html"|"text"}.

    drift = Phase 0 + drift-only tables (MAAS-only, matched with drift, NIC drift, OS gaps).
    reference = full matched hosts, full per-device NIC audit, OpenStack ref (collapsible in UI).
    OpenStack data is already combined from all configured clouds before being passed here.

    When use_html is True (default), drift is a safe HTML fragment for |safe in templates.
    """
    orphaned_nb_count = len((drift or {}).get("in_netbox_not_maas") or [])
    drift = _drift_for_user_reports(drift)
    e = _DriftReportEmitter(use_html=use_html)
    ref_lines = []

    # --- INVENTORY (compact) ---
    e.banner("INVENTORY")
    e.spacer()
    e.subtitle("Run metadata")
    e.spacer()
    e.table(
        ["Property", "Value"],
        _run_metadata_rows(maas_data, netbox_data, openstack_data),
        dynamic_columns=True,
    )
    e.spacer()
    e.subtitle("MAAS")
    e.spacer()
    if maas_data.get("error"):
        e.error(f"Error: {maas_data['error']}")
    else:
        e.table(
            ["Metric", "Count"],
            [
                ["Zones", str(len(maas_data.get("zones") or []))],
                ["Resource pools", str(len(maas_data.get("pools") or []))],
                ["Machines", str(len(maas_data.get("machines") or []))],
            ],
        )

    e.spacer()
    e.subtitle("NetBox (this instance)")
    e.spacer()
    if netbox_data.get("error"):
        e.error(f"Error: {netbox_data['error']}")
    else:
        inv_rows = [
            ["Sites", str(len(netbox_data.get("sites") or []))],
            ["Devices", str(len(netbox_data.get("devices") or []))],
        ]
        if netbox_prefix_count:
            inv_rows.append(["IPAM Prefix objects", str(netbox_prefix_count)])
        e.table(["Metric", "Count"], inv_rows)

    scope_meta = (drift or {}).get("scope_meta") or {}
    if scope_meta:
        e.spacer()
        e.banner("SCOPE", "-")
        e.spacer()
        sel_sites = ", ".join(scope_meta.get("selected_sites") or []) or "(all)"
        sel_locs = ", ".join(scope_meta.get("selected_locations") or []) or "(all)"
        e.table(
            ["Check", "Value"],
            [
                ["Coverage status", str(scope_meta.get("coverage_status") or "PARTIAL")],
                ["Selected sites", sel_sites],
                ["Selected locations", sel_locs],
                [
                    "MAAS machines included / fetched",
                    f"{scope_meta.get('maas_machines_after', 0)} / {scope_meta.get('maas_machines_before', 0)}",
                ],
                [
                    "NetBox devices included / fetched",
                    f"{scope_meta.get('netbox_devices_after', 0)} / {scope_meta.get('netbox_devices_before', 0)}",
                ],
                [
                    "OpenStack nets included / fetched",
                    f"{scope_meta.get('openstack_networks_after', 0)} / {scope_meta.get('openstack_networks_before', 0)}",
                ],
                [
                    "OpenStack subnets included / fetched",
                    f"{scope_meta.get('openstack_subnets_after', 0)} / {scope_meta.get('openstack_subnets_before', 0)}",
                ],
                [
                    "OpenStack FIPs included / fetched",
                    f"{scope_meta.get('openstack_fips_after', 0)} / {scope_meta.get('openstack_fips_before', 0)}",
                ],
            ],
            dynamic_columns=True,
        )

    # --- Phase 0 — drift category counts ---
    e.spacer()
    e.banner("DRIFT COUNTS")
    e.paragraph("Counts for this run (match by hostname and NIC MAC).")
    e.spacer()
    pc = _phase0_category_counts(
        drift,
        matched_rows,
        interface_audit,
        os_subnet_gaps,
        os_floating_gaps,
    )
    serial_validation_needed = _count_hints(matched_rows, "NB serial empty")
    bmc_oob_mismatch = _count_hints(matched_rows, "MAAS BMC ")
    sub_txt = str(pc["sub_gaps"]) if pc["sub_gaps"] is not None else "N/A (local ORM)"
    e.table(
        ["Category", "Count"],
        [
            ["In MAAS only (not in NetBox)", str(pc["maas_only"])],
            [
                "Orphaned NetBox devices (not seen in MAAS this run; read-only here; "
                "tagging/cleanup deferred to a separate UI workflow because NetBox update sources include netbox-agent, scripts, and manual entries, not just MAAS)",
                str(orphaned_nb_count),
            ],
            ["Matched — review hints", str(pc["check_hosts"])],
            ["NetBox serial missing", str(serial_validation_needed)],
            ["NIC rows not OK", str(pc["iface_not_ok"])],
            ["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])],
            ["VLAN mismatch (MAAS vs NetBox)", str(pc["vlan_drift_nic"])],
            ["VLAN unverified from MAAS", str(pc["vlan_unverified_nic"])],
            ["OpenStack subnet → no Prefix", sub_txt],
            ["OpenStack FIP → no IP record", str(pc["fip_gaps"])],
            ["BMC vs NetBox OOB differs", str(bmc_oob_mismatch)],
            ["LLDP / cabling", "—"],
        ],
    )

    # --- Severity triage ---
    e.spacer()
    e.banner("SEVERITY TRIAGE (why these matter)", "-")
    e.paragraph("Priority rules used in this report for review ordering.")
    e.spacer()
    sev_rows = _severity_triage_rows(
        pc,
        serial_validation_needed=serial_validation_needed,
        bmc_oob_mismatch=bmc_oob_mismatch,
    )
    e.table(
        ["Severity", "Category", "Count", "Why this matters"],
        sev_rows,
        wrap_max_width=None,
    )

    # --- Run metrics ---
    e.spacer()
    e.banner("RUN METRICS", "-")
    e.spacer()
    e.table(
        ["Metric", "Value"],
        [
            ["MAAS machines", str(len(maas_data.get("machines") or []))],
            ["NetBox devices", str(len(netbox_data.get("devices") or []))],
            ["Matched hostnames", str(drift.get("matched_count", 0))],
            ["In MAAS only", str(pc["maas_only"])],
            ["NetBox serial missing", str(serial_validation_needed)],
            ["OpenStack subnet gaps", sub_txt],
            ["OpenStack FIP gaps", str(pc["fip_gaps"])],
            ["VLAN mismatch NICs", str(pc["vlan_drift_nic"])],
            ["VLAN unverified NICs", str(pc["vlan_unverified_nic"])],
            ["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])],
        ],
    )

    align_rows = _alignment_review_rows(matched_rows)
    if align_rows:
        e.spacer()
        e.subtitle("Detail — placement & lifecycle alignment")
        e.spacer()
        e.table(
            [
                "Host",
                "MAAS fabric",
                "NetBox site",
                "NetBox location",
                "MAAS state",
                "NB state",
                "Alignment issues",
            ],
            align_rows,
            dynamic_columns=True,
            notes_col_idx=6,
            wrap_max_width=None,
        )

    # --- Proposed changes (preview only; full list, uncapped) ---
    prop = _proposed_changes_rows(
        maas_data,
        netbox_data,
        drift,
        interface_audit,
        matched_rows,
        os_subnet_gaps or [],
        os_floating_gaps or [],
        netbox_ifaces=netbox_ifaces,
    )
    e.spacer()
    e.banner("PROPOSED CHANGES", "-")
    e.paragraph(
        "Read-only. Possible NetBox updates from MAAS and OpenStack — nothing is applied from this screen."
    )
    e.spacer()

    e.subtitle("A) Add to NetBox")
    e.spacer()
    e.table(
        ["What", "Count", "Note"],
        [
            ["New devices (MAAS)", str(len(prop["add_devices"])), "Safe create candidates"],
            [
                "Review-only MAAS-only hosts",
                str(len(prop.get("add_devices_review_only", []))),
                "Not safe to auto-propose (status/data quality policy)",
            ],
            ["New prefixes (OpenStack)", str(len(prop["add_prefixes"])), "Subnet not in IPAM"],
            ["New floating IPs (OpenStack)", str(len(prop["add_fips"])), "FIP not in IPAM"],
        ],
    )
    if prop["add_devices"]:
        e.spacer()
        e.subtitle("Detail — new devices")
        e.spacer()
        e.table(
            [
                "Hostname",
                "NB region",
                "NB site",
                "NB location",
                "NetBox device type",
                "NetBox role",
                "MAAS fabric",
                "MAAS status",
                "Serial Number",
                "Power type",
                "BMC present",
                "NIC count",
                "Primary MAC (MAAS)",
                "Proposed Tag",
                "Proposed Action",
            ],
            prop["add_devices"],
            dynamic_columns=True,
            wrap_max_width=None,
        )
    if prop.get("add_devices_review_only"):
        e.spacer()
        e.subtitle("Detail — MAAS-only review-only (not safe add candidates)")
        e.spacer()
        e.table(
            [
                "Hostname",
                "NB region",
                "NB site",
                "NB location",
                "NetBox device type",
                "NetBox role",
                "MAAS fabric",
                "MAAS status",
                "Serial Number",
                "Power type",
                "BMC present",
                "NIC count",
                "Primary MAC (MAAS)",
                "Proposed Tag",
                "Proposed Action",
            ],
            prop["add_devices_review_only"],
            dynamic_columns=True,
            wrap_max_width=None,
        )
    if prop["add_prefixes"]:
        e.spacer()
        e.subtitle("Detail — new prefixes")
        e.spacer()
        e.table(
            ["CIDR", "Network Name", "Network ID", "Cloud", "Proposed Action"],
            prop["add_prefixes"],
            dynamic_columns=True,
            wrap_max_width=None,
        )
    if prop["add_fips"]:
        e.spacer()
        e.subtitle("Detail — new floating IPs")
        e.spacer()
        e.table(
            ["Floating IP", "Fixed IP", "Project", "Cloud", "Proposed Action"],
            prop["add_fips"],
            dynamic_columns=True,
            wrap_max_width=None,
        )

    e.spacer()
    e.subtitle("B) NICs and BMC / OOB")
    e.spacer()
    e.table(
        ["What", "Count", "Note"],
        [
            ["New NICs in NetBox", str(len(prop["add_nb_interfaces"])), "MAAS MAC not on device"],
            ["NIC drift", str(len(prop["update_nic"])), "MAAS vs NetBox differs"],
            [
                "BMC / OOB",
                str(len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", []))),
                "Power / out-of-band vs NetBox",
            ],
        ],
    )
    if prop["add_nb_interfaces"]:
        e.spacer()
        e.subtitle("Detail — new NICs")
        e.spacer()
        e.table(
            [
                "Host",
                "NB site",
                "NB location",
                "MAAS intf",
                "MAAS fabric",
                "MAAS MAC",
                "MAAS IPs",
                "MAAS VLAN",
                "Suggested NB name",
                "Proposed properties (from MAAS)",
                "Risk",
            ],
            prop["add_nb_interfaces"],
            dynamic_columns=True,
            wrap_max_width=None,
        )
    if prop["update_nic"]:
        e.spacer()
        e.subtitle("Detail — NIC drift")
        e.spacer()
        e.table(
            [
                "Host",
                "MAAS intf",
                "MAAS fabric",
                "MAAS MAC",
                "MAAS IPs",
                "NB intf",
                "NB MAC",
                "NB IPs",
                "MAAS VLAN",
                "NB VLAN",
                "Status",
                "Reason",
                "Proposed Action",
                "Risk",
            ],
            prop["update_nic"],
            dynamic_columns=True,
            wrap_max_width=None,
        )

    if prop.get("add_mgmt_iface_new_devices"):
        e.spacer()
        e.subtitle("Detail — new BMC / OOB interfaces")
        e.spacer()
        e.table(
            [
                "Host",
                "MAAS BMC IP",
                "MAAS power_type",
                "MAAS BMC MAC",
                "Suggested NB mgmt iface",
                "NB mgmt iface IP",
                "Proposed action",
                "Risk",
            ],
            prop["add_mgmt_iface_new_devices"],
            dynamic_columns=True,
            wrap_max_width=None,
        )

    if prop["add_mgmt_iface"]:
        e.spacer()
        e.subtitle("Detail — BMC / OOB")
        e.spacer()
        e.table(
            [
                "Host",
                "MAAS BMC IP",
                "MAAS power_type",
                "MAAS BMC MAC",
                "NB OOB port (hint)",
                "NetBox OOB",
                "NB IP coverage",
                "NB port w/ BMC IP",
                "NB OOB MAC",
                "Status",
                "Proposed action",
                "Risk",
            ],
            prop["add_mgmt_iface"],
            dynamic_columns=True,
            wrap_max_width=None,
        )

    e.spacer()
    e.subtitle("C) Review")
    e.spacer()
    e.table(
        ["What", "Count", "Note"],
        [
            ["Serial check", str(len(prop["review_serial"])), "NetBox serial empty"],
        ],
    )
    if prop["review_serial"]:
        e.spacer()
        e.subtitle("Detail — serials")
        e.spacer()
        e.table(
            ["Hostname", "MAAS Serial", "NetBox Serial", "Proposed Action", "Risk"],
            prop["review_serial"],
            dynamic_columns=True,
            wrap_max_width=None,
        )
    e.spacer()
    e.subtitle("Summary")
    e.spacer()
    total_props = (
        len(prop["add_devices"]) + len(prop["add_prefixes"]) + len(prop["add_fips"]) +
        len(prop["update_nic"]) + len(prop["add_nb_interfaces"]) +
        len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", [])) +
        len(prop["review_serial"])
    )
    e.table(
        ["Bucket", "Count"],
        [
            ["New devices", str(len(prop["add_devices"]))],
            ["New prefixes", str(len(prop["add_prefixes"]))],
            ["New floating IPs", str(len(prop["add_fips"]))],
            ["NIC drift", str(len(prop["update_nic"]))],
            ["New NICs", str(len(prop["add_nb_interfaces"]))],
            ["BMC / OOB", str(len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", [])))],
            ["Serials (review)", str(len(prop["review_serial"]))],
            ["Total", str(total_props)],
        ],
    )

    e.spacer()
    e.banner("END OF DRIFT AUDIT", "=")

    # ---------- REFERENCE ----------
    # Hidden intentionally for user-facing output.

    return {
        "drift": e.render(),
        "reference": "\n".join(ref_lines),
        "drift_markup": "html" if use_html else "text",
    }


def build_drift_report_xlsx(
    maas_data,
    netbox_data,
    openstack_data,
    drift,
    *,
    matched_rows=None,
    os_subnet_hints=None,
    os_subnet_gaps=None,
    os_floating_gaps=None,
    netbox_prefix_count=0,
    interface_audit=None,
    netbox_ifaces=None,
):
    """
    Build an Excel (.xlsx) workbook from the same inputs as format_drift_report.
    Returns bytes suitable for HttpResponse(..., content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet").
    Google Sheets opens .xlsx files.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export. pip install openpyxl")

    orphaned_nb_count = len((drift or {}).get("in_netbox_not_maas") or [])
    drift = _drift_for_user_reports(drift)

    wb = Workbook()
    header_font = Font(bold=True)

    def _sheet(name, max_len=31):
        s = wb.create_sheet(title=name[:max_len])
        return s

    def _append_header(ws, row):
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(row) + 1):
            ws.cell(row=r, column=c).font = header_font
        return r

    # --- Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Drift audit summary"])
    ws_sum.cell(row=1, column=1).font = header_font
    ws_sum.append([])
    r_own = ws_sum.max_row + 1
    ws_sum.append([_PHASE0_FIELD_OWNERSHIP_TITLE, "", ""])
    ws_sum.cell(row=r_own, column=1).font = header_font
    ws_sum.append([_PHASE0_FIELD_OWNERSHIP_LEAD, "", ""])
    for b in _PHASE0_FIELD_OWNERSHIP_BULLETS:
        ws_sum.append([f"  • {b}", "", ""])
    ws_sum.append([])
    r_rm = ws_sum.max_row + 1
    ws_sum.append(["RUN METADATA", ""])
    ws_sum.cell(row=r_rm, column=1).font = header_font
    _append_header(ws_sum, ["Property", "Value"])
    for prop, val in _run_metadata_rows(maas_data, netbox_data, openstack_data):
        ws_sum.append([prop, val])
    ws_sum.append([])
    ws_sum.append(["MAAS", "OK" if not maas_data.get("error") else "Error", ""])
    ws_sum.append(["  Machines", str(len(maas_data.get("machines") or [])), ""])
    ws_sum.append(["NetBox", "OK" if not netbox_data.get("error") else "Error", ""])
    ws_sum.append(["  Devices", str(len(netbox_data.get("devices") or [])), ""])
    ws_sum.append(["  Sites", str(len(netbox_data.get("sites") or [])), ""])
    if netbox_prefix_count:
        ws_sum.append(["  IPAM Prefixes", str(netbox_prefix_count), ""])
    ws_sum.append([])
    ws_sum.append([])
    pc = _phase0_category_counts(
        drift,
        matched_rows,
        interface_audit,
        os_subnet_gaps,
        os_floating_gaps or [],
    )
    serial_validation_needed = _count_hints(matched_rows, "NB serial empty")
    bmc_oob_mismatch = _count_hints(matched_rows, "MAAS BMC ")
    sub_txt = str(pc["sub_gaps"]) if pc["sub_gaps"] is not None else "N/A"
    scope_meta = (drift or {}).get("scope_meta") or {}
    if scope_meta:
        ws_sum.append([])
        ws_sum.append(["SCOPE", "", ""])
        _append_header(ws_sum, ["Check", "Value"])
        ws_sum.append(["Coverage status", str(scope_meta.get("coverage_status") or "PARTIAL")])
        ws_sum.append(["Selected sites", ", ".join(scope_meta.get("selected_sites") or []) or "(all)"])
        ws_sum.append(["Selected locations", ", ".join(scope_meta.get("selected_locations") or []) or "(all)"])
        ws_sum.append([
            "MAAS machines included / fetched",
            f"{scope_meta.get('maas_machines_after', 0)} / {scope_meta.get('maas_machines_before', 0)}",
        ])
        ws_sum.append([
            "NetBox devices included / fetched",
            f"{scope_meta.get('netbox_devices_after', 0)} / {scope_meta.get('netbox_devices_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack nets included / fetched",
            f"{scope_meta.get('openstack_networks_after', 0)} / {scope_meta.get('openstack_networks_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack subnets included / fetched",
            f"{scope_meta.get('openstack_subnets_after', 0)} / {scope_meta.get('openstack_subnets_before', 0)}",
        ])
        ws_sum.append([
            "OpenStack FIPs included / fetched",
            f"{scope_meta.get('openstack_fips_after', 0)} / {scope_meta.get('openstack_fips_before', 0)}",
        ])
    ws_sum.append([])
    ws_sum.append(["DRIFT COUNTS", "", ""])
    _append_header(ws_sum, ["Category", "Count"])
    ws_sum.append(["In MAAS only (not in NetBox)", str(pc["maas_only"])])
    ws_sum.append(
        [
            "Orphaned NetBox devices (not seen in MAAS this run; read-only here; tagging/cleanup deferred to a separate UI workflow because NetBox update sources include netbox-agent, scripts, and manual entries, not just MAAS)",
            str(orphaned_nb_count),
        ]
    )
    ws_sum.append(["Matched — review hints", str(pc["check_hosts"])])
    ws_sum.append(["NetBox serial missing", str(serial_validation_needed)])
    ws_sum.append(["NIC rows not OK", str(pc["iface_not_ok"])])
    ws_sum.append(["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])])
    ws_sum.append(["VLAN mismatch (MAAS vs NetBox)", str(pc["vlan_drift_nic"])])
    ws_sum.append(["VLAN unverified from MAAS", str(pc["vlan_unverified_nic"])])
    ws_sum.append(["OpenStack subnet → no Prefix", sub_txt])
    ws_sum.append(["OpenStack FIP → no IP record", str(pc["fip_gaps"])])
    ws_sum.append(["BMC vs NetBox OOB differs", str(bmc_oob_mismatch)])
    ws_sum.append(["LLDP / cabling", "—"])
    ws_sum.append([])
    ws_sum.append(["SEVERITY TRIAGE (why these matter)", "", ""])
    _append_header(ws_sum, ["Severity", "Category", "Count", "Why this matters"])
    for row in _severity_triage_rows(
        pc,
        serial_validation_needed=serial_validation_needed,
        bmc_oob_mismatch=bmc_oob_mismatch,
    ):
        ws_sum.append(row)
    ws_sum.append([])
    ws_sum.append(["RUN METRICS", "", ""])
    _append_header(ws_sum, ["Metric", "Value"])
    ws_sum.append(["MAAS machines", str(len(maas_data.get("machines") or []))])
    ws_sum.append(["NetBox devices", str(len(netbox_data.get("devices") or []))])
    ws_sum.append(["Matched hostnames", str(drift.get("matched_count", 0))])
    ws_sum.append(["In MAAS only", str(pc["maas_only"])])
    ws_sum.append(["NetBox serial missing", str(serial_validation_needed)])
    ws_sum.append(["OpenStack subnet gaps", sub_txt])
    ws_sum.append(["OpenStack FIP gaps", str(pc["fip_gaps"])])
    ws_sum.append(["VLAN mismatch NICs", str(pc["vlan_drift_nic"])])
    ws_sum.append(["VLAN unverified NICs", str(pc["vlan_unverified_nic"])])
    ws_sum.append(["MAAS NIC missing in NetBox", str(pc["maas_nic_missing_nb"])])
    ws_sum.append([])
    align_rows_x = _alignment_review_rows(matched_rows)
    if align_rows_x:
        r_al = ws_sum.max_row + 1
        ws_sum.append(["Detail — placement & lifecycle alignment", ""])
        ws_sum.cell(row=r_al, column=1).font = header_font
        _append_header(
            ws_sum,
            [
                "Host",
                "MAAS fabric",
                "NetBox site",
                "NetBox location",
                "MAAS state",
                "NB state",
                "Alignment issues",
            ],
        )
        for row in align_rows_x:
            ws_sum.append(row)
        ws_sum.append([])

    prop = _proposed_changes_rows(
        maas_data,
        netbox_data,
        drift,
        interface_audit,
        matched_rows,
        os_subnet_gaps or [],
        os_floating_gaps or [],
        netbox_ifaces=netbox_ifaces,
    )
    ws_sum.append([])
    ws_sum.append(["PROPOSED CHANGES (read-only)", "", ""])
    _append_header(ws_sum, ["Bucket", "Count"])
    total_props_x = (
        len(prop["add_devices"])
        + len(prop["add_prefixes"])
        + len(prop["add_fips"])
        + len(prop["update_nic"])
        + len(prop["add_nb_interfaces"])
        + len(prop["add_mgmt_iface"])
        + len(prop.get("add_mgmt_iface_new_devices", []))
        + len(prop["review_serial"])
    )
    ws_sum.append(["New devices", str(len(prop["add_devices"]))])
    ws_sum.append(["Review-only MAAS-only hosts", str(len(prop.get("add_devices_review_only", [])))])
    ws_sum.append(["New prefixes", str(len(prop["add_prefixes"]))])
    ws_sum.append(["New floating IPs", str(len(prop["add_fips"]))])
    ws_sum.append(["NIC drift", str(len(prop["update_nic"]))])
    ws_sum.append(["New NICs", str(len(prop["add_nb_interfaces"]))])
    ws_sum.append(["BMC / OOB", str(len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", [])))])
    ws_sum.append(["Serials (review)", str(len(prop["review_serial"]))])
    ws_sum.append(["Total", str(total_props_x)])

    # Matched-host drift worksheet intentionally suppressed to match on-screen report.

    # --- Proposed changes (full list) ---
    ws_prop = _sheet("Proposed changes")
    ws_prop.append(["Drift detail — read-only; nothing is written to NetBox from this export."])
    ws_prop.cell(row=1, column=1).font = header_font
    ws_prop.append([])
    _append_header(ws_prop, ["Section", "Count"])
    ws_prop.append(["New devices (MAAS)", len(prop["add_devices"])])
    ws_prop.append(["Review-only MAAS-only hosts", len(prop.get("add_devices_review_only", []))])
    ws_prop.append(["New prefixes (OpenStack)", len(prop["add_prefixes"])])
    ws_prop.append(["New floating IPs (OpenStack)", len(prop["add_fips"])])
    ws_prop.append(["NIC drift", len(prop["update_nic"])])
    ws_prop.append(["New NICs", len(prop["add_nb_interfaces"])])
    ws_prop.append(["BMC / OOB", len(prop["add_mgmt_iface"]) + len(prop.get("add_mgmt_iface_new_devices", []))])
    ws_prop.append(["Serials (review)", len(prop["review_serial"])])

    def _append_block(title, headers, rows):
        ws_prop.append([])
        ws_prop.append([title])
        ws_prop.cell(row=ws_prop.max_row, column=1).font = header_font
        _append_header(ws_prop, headers)
        for row in rows:
            ws_prop.append(list(row))

    _append_block(
        "A) New devices",
        [
            "Hostname",
            "NB region",
            "NB site",
            "NB location",
            "NetBox device type",
            "NetBox role",
            "MAAS fabric",
            "MAAS status",
            "Serial Number",
            "Power type",
            "BMC present",
            "NIC count",
            "Primary MAC (MAAS)",
            "Proposed Tag",
            "Proposed Action",
        ],
        prop["add_devices"],
    )
    _append_block(
        "A) MAAS-only review-only (not safe add candidates)",
        [
            "Hostname",
            "NB region",
            "NB site",
            "NB location",
            "NetBox device type",
            "NetBox role",
            "MAAS fabric",
            "MAAS status",
            "Serial Number",
            "Power type",
            "BMC present",
            "NIC count",
            "Primary MAC (MAAS)",
            "Proposed Tag",
            "Proposed Action",
        ],
        prop.get("add_devices_review_only", []),
    )
    _append_block(
        "A) New prefixes",
        ["CIDR", "Network Name", "Network ID", "Cloud", "Proposed Action"],
        prop["add_prefixes"],
    )
    _append_block(
        "A) New floating IPs",
        ["Floating IP", "Fixed IP", "Project", "Cloud", "Proposed Action"],
        prop["add_fips"],
    )
    _append_block(
        "B) New NICs",
        [
            "Host",
            "NB site",
            "NB location",
            "MAAS intf",
            "MAAS fabric",
            "MAAS MAC",
            "MAAS IPs",
            "MAAS VLAN",
            "Suggested NB name",
            "Proposed properties (from MAAS)",
            "Risk",
        ],
        prop["add_nb_interfaces"],
    )
    _append_block(
        "B) NIC drift",
        [
            "Host",
            "MAAS intf",
            "MAAS fabric",
            "MAAS MAC",
            "MAAS IPs",
            "NB intf",
            "NB MAC",
            "NB IPs",
            "MAAS VLAN",
            "NB VLAN",
            "Status",
            "Reason",
            "Proposed Action",
            "Risk",
        ],
        prop["update_nic"],
    )
    _append_block(
        "B) BMC / OOB",
        [
            "Host",
            "MAAS BMC IP",
            "MAAS power_type",
            "MAAS BMC MAC",
            "NB OOB port (hint)",
            "NetBox OOB",
            "NB IP coverage",
            "NB port w/ BMC IP",
            "NB OOB MAC",
            "Status",
            "Proposed action",
            "Risk",
        ],
        prop["add_mgmt_iface"],
    )
    _append_block(
        "B) New-device BMC / OOB interfaces",
        [
            "Host",
            "MAAS BMC IP",
            "MAAS power_type",
            "MAAS BMC MAC",
            "Suggested NB mgmt iface",
            "NB mgmt iface IP",
            "Proposed action",
            "Risk",
        ],
        prop.get("add_mgmt_iface_new_devices", []),
    )
    _append_block(
        "C) Serials",
        ["Hostname", "MAAS Serial", "NetBox Serial", "Proposed Action", "Risk"],
        prop["review_serial"],
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
